from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from sprint_cert_automation.domain.models import Holiday, SprintWindow, TeamMember, TeamMemberWorkload
from sprint_cert_automation.domain.rules import free_hours_from_non_working_days, sprint_hours
from sprint_cert_automation.utils.dates import previous_month, year_month_label

TEAM_ROWS = {
    1: "Bonificaciones",
    2: "Subvenciones",
    3: "Fondos de Reserva MRR",
    4: "AccentureTransversal",
}

MONTH_TOKENS = {
    1: "ene",
    2: "feb",
    3: "mar",
    4: "abr",
    5: "mayo",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "sept",
    10: "oct",
    11: "nov",
    12: "dic",
}

SPRINT_ID_PATTERN = re.compile(r"SP\d+", re.IGNORECASE)
DEFAULT_DAY_HOURS = 8.5
HEADER_ROW = 6
CALENDAR_ROW = 5
DATA_START_ROW = 7
NON_WORKING_FILL_COLORS = {13684944}
HOLIDAY_FILL_RGB = {"FFFF00", "FF0000"}
NON_WORKING_THEME = 2
NON_WORKING_TINT = -0.0999786370433668
TEAM_ALIASES = {
    "Bonificaciones": "Bonificaciones",
    "Subvenciones": "Subvenciones",
    "Fondos de Reserva": "Fondos de Reserva MRR",
    "Fondos de Reserva MRR": "Fondos de Reserva MRR",
    "Transversal": "AccentureTransversal",
    "Transversales": "AccentureTransversal",
    "AccentureTransversal": "AccentureTransversal",
}


@dataclass(frozen=True)
class _SheetWindow:
    year: int
    month: int
    name: str


@dataclass(frozen=True)
class _SprintSegment:
    team: str
    sprint_id: str
    start_date: date
    end_date: date
    source_sheet: str
    is_hatched: bool


@dataclass(frozen=True)
class _SheetContext:
    worksheet: Worksheet
    window: _SheetWindow
    day_columns: dict[int, int]
    headers: dict[str, int]
    member_rows: dict[str, int]


class ForecastReader:
    """Read forecast workbook and return normalized domain objects."""

    def __init__(self, forecast_path: Path) -> None:
        self.forecast_path = forecast_path

    def read_sprints_for_target_window(self, year: int, month: int) -> list[SprintWindow]:
        workbook = load_workbook(self.forecast_path, data_only=False)
        try:
            windows = self._sheet_windows(year, month)
            segments: list[_SprintSegment] = []
            for window in windows:
                if window.name not in workbook.sheetnames:
                    raise ValueError(f"Forecast sheet not found: {window.name}")
                worksheet = workbook[window.name]
                segments.extend(self._read_sheet_segments(worksheet, window))
            return self._merge_segments(segments, year, month)
        finally:
            workbook.close()

    def read_draft_data_for_sprint(
        self,
        sprint: SprintWindow,
    ) -> tuple[list[Holiday], list[TeamMemberWorkload]]:
        workbook = load_workbook(self.forecast_path, data_only=False)
        try:
            windows = self._sprint_windows(sprint)
            contexts = [self._sheet_context(workbook, window) for window in windows]
            holidays = self._read_holidays_for_sprint(contexts, sprint)
            workloads = self._read_workloads_for_sprint(contexts, sprint, holidays)
            return holidays, workloads
        finally:
            workbook.close()

    def _sheet_windows(self, year: int, month: int) -> list[_SheetWindow]:
        previous_year, previous_target_month = previous_month(year, month)
        return [
            _SheetWindow(
                year=previous_year,
                month=previous_target_month,
                name=self._sheet_name(previous_year, previous_target_month),
            ),
            _SheetWindow(
                year=year,
                month=month,
                name=self._sheet_name(year, month),
            ),
        ]

    def _sheet_name(self, year: int, month: int) -> str:
        fiscal_year = year if month <= 8 else year + 1
        return f"FY{fiscal_year % 100:02d}_{MONTH_TOKENS[month]}"

    def _sprint_windows(self, sprint: SprintWindow) -> list[_SheetWindow]:
        windows: list[_SheetWindow] = []
        seen: set[str] = set()
        for year, month in {
            (sprint.start_date.year, sprint.start_date.month),
            (sprint.end_date.year, sprint.end_date.month),
        }:
            window = _SheetWindow(year=year, month=month, name=self._sheet_name(year, month))
            if window.name not in seen:
                windows.append(window)
                seen.add(window.name)
        windows.sort(key=lambda item: (item.year, item.month))
        return windows

    def _sheet_context(self, workbook, window: _SheetWindow) -> _SheetContext:
        if window.name not in workbook.sheetnames:
            raise ValueError(f"Forecast sheet not found: {window.name}")
        worksheet = workbook[window.name]
        headers = self._header_columns(worksheet)
        member_rows = self._member_rows(worksheet, headers)
        return _SheetContext(
            worksheet=worksheet,
            window=window,
            day_columns=self._day_columns(worksheet),
            headers=headers,
            member_rows=member_rows,
        )

    def _read_sheet_segments(
        self,
        worksheet: Worksheet,
        window: _SheetWindow,
    ) -> list[_SprintSegment]:
        day_columns = self._day_columns(worksheet)
        if not day_columns:
            return []
        first_day_column = min(day_columns)
        last_day_column = max(day_columns)

        covered_cells: set[tuple[int, int]] = set()
        segments: list[_SprintSegment] = []

        for row_index, team in TEAM_ROWS.items():
            for merged_range in worksheet.merged_cells.ranges:
                if merged_range.min_row != row_index or merged_range.max_row != row_index:
                    continue
                if merged_range.max_col < first_day_column or merged_range.min_col > last_day_column:
                    continue

                start_col = max(merged_range.min_col, first_day_column)
                end_col = min(merged_range.max_col, last_day_column)
                top_left = worksheet.cell(row=row_index, column=merged_range.min_col)
                segment = self._build_segment(
                    team=team,
                    header_text=str(top_left.value or "").strip(),
                    fill=top_left.fill,
                    start_day=day_columns[start_col],
                    end_day=day_columns[end_col],
                    sheet_name=window.name,
                    year=window.year,
                    month=window.month,
                )
                if segment is not None:
                    segments.append(segment)
                for column_index in range(start_col, end_col + 1):
                    covered_cells.add((row_index, column_index))

            for column_index, day_value in day_columns.items():
                if (row_index, column_index) in covered_cells:
                    continue
                cell = worksheet.cell(row=row_index, column=column_index)
                header_text = str(cell.value or "").strip()
                if not header_text:
                    continue
                segment = self._build_segment(
                    team=team,
                    header_text=header_text,
                    fill=cell.fill,
                    start_day=day_value,
                    end_day=day_value,
                    sheet_name=window.name,
                    year=window.year,
                    month=window.month,
                )
                if segment is not None:
                    segments.append(segment)

        return segments

    def _day_columns(self, worksheet: Worksheet) -> dict[int, int]:
        day_columns: dict[int, int] = {}
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=CALENDAR_ROW, column=column_index).value
            if isinstance(value, int) and 1 <= value <= 31:
                day_columns[column_index] = value
            elif isinstance(value, float) and value.is_integer() and 1 <= int(value) <= 31:
                day_columns[column_index] = int(value)
        return day_columns

    def _header_columns(self, worksheet: Worksheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=HEADER_ROW, column=column_index).value
            if isinstance(value, str) and value.strip():
                headers[value.strip()] = column_index
        return headers

    def _member_rows(self, worksheet: Worksheet, headers: dict[str, int]) -> dict[str, int]:
        technician_column = self._technician_column(headers, required=False)
        if technician_column is None:
            return {}

        rows: dict[str, int] = {}
        for row_index in range(DATA_START_ROW, worksheet.max_row + 1):
            display_name = self._string_value(worksheet.cell(row=row_index, column=technician_column).value)
            if not display_name:
                continue
            if display_name not in rows:
                rows[display_name] = row_index
        return rows

    def _build_segment(
        self,
        team: str,
        header_text: str,
        fill: PatternFill,
        start_day: int,
        end_day: int,
        sheet_name: str,
        year: int,
        month: int,
    ) -> _SprintSegment | None:
        sprint_id = self._sprint_id_from_header(team, header_text, year, month)
        if sprint_id is None:
            return None
        return _SprintSegment(
            team=team,
            sprint_id=sprint_id,
            start_date=date(year, month, start_day),
            end_date=date(year, month, end_day),
            source_sheet=sheet_name,
            is_hatched=self._is_hatched_fill(fill),
        )

    def _sprint_id_from_header(
        self,
        team: str,
        header_text: str,
        year: int,
        month: int,
    ) -> str | None:
        if not header_text:
            return None
        if team == "AccentureTransversal":
            if "transversal" not in header_text.lower():
                return None
            return year_month_label(year, month)

        match = SPRINT_ID_PATTERN.search(header_text)
        if match is None:
            return None
        return match.group(0).upper()

    def _is_hatched_fill(self, fill: PatternFill) -> bool:
        pattern_type = fill.patternType
        return pattern_type not in (None, "solid")

    def _read_holidays_for_sprint(
        self,
        contexts: list[_SheetContext],
        sprint: SprintWindow,
    ) -> list[Holiday]:
        holidays: dict[date, Holiday] = {}
        for context in contexts:
            for column_index, day in context.day_columns.items():
                holiday_date = date(context.window.year, context.window.month, day)
                if holiday_date < sprint.start_date or holiday_date > sprint.end_date:
                    continue
                cell = context.worksheet.cell(row=CALENDAR_ROW, column=column_index)
                if self._is_holiday_fill(cell.fill):
                    holidays[holiday_date] = Holiday(holiday_date=holiday_date, label="Festivo")
        return [holidays[key] for key in sorted(holidays)]

    def _read_workloads_for_sprint(
        self,
        contexts: list[_SheetContext],
        sprint: SprintWindow,
        holidays: list[Holiday],
    ) -> list[TeamMemberWorkload]:
        target_context = contexts[-1]
        team_members = self._team_members(target_context, sprint.team)
        holiday_dates = {holiday.holiday_date for holiday in holidays}
        total_sprint_hours = self._sprint_hours_between(sprint.start_date, sprint.end_date, holiday_dates)

        workloads: list[TeamMemberWorkload] = []
        for member in team_members:
            non_working_days = 0
            for context in contexts:
                row_index = context.member_rows.get(member.name)
                if row_index is None:
                    continue
                for column_index in self._sprint_columns(context, sprint):
                    day_date = date(context.window.year, context.window.month, context.day_columns[column_index])
                    if day_date in holiday_dates:
                        continue
                    cell = context.worksheet.cell(row=row_index, column=column_index)
                    if self._is_non_working_fill(cell.fill):
                        non_working_days += 1
            workloads.append(
                TeamMemberWorkload(
                    member=member,
                    sprint_hours=total_sprint_hours,
                    free_hours=free_hours_from_non_working_days(non_working_days, DEFAULT_DAY_HOURS),
                )
            )

        return workloads

    def _team_members(self, context: _SheetContext, team: str) -> list[TeamMember]:
        technician_column = self._technician_column(context.headers, required=True)
        team_column = context.headers.get("Equipo")
        billing_column = context.headers.get("Facturación")
        category_column = context.headers.get("Perfil Facturable")
        if team_column is None or category_column is None:
            raise ValueError(f"Required team columns not found in sheet {context.worksheet.title}")

        members: list[TeamMember] = []
        seen_member_names: set[str] = set()
        for row_index in range(DATA_START_ROW, context.worksheet.max_row + 1):
            member_team = self._normalize_team_name(
                self._string_value(context.worksheet.cell(row=row_index, column=team_column).value)
            )
            if member_team != team:
                continue
            name = self._string_value(context.worksheet.cell(row=row_index, column=technician_column).value)
            if not name:
                continue
            normalized_name = " ".join(name.split()).casefold()
            if normalized_name in seen_member_names:
                continue
            seen_member_names.add(normalized_name)
            billing_line = ""
            if billing_column is not None:
                billing_line = self._string_value(
                    context.worksheet.cell(row=row_index, column=billing_column).value
                )
            category = self._string_value(context.worksheet.cell(row=row_index, column=category_column).value)
            members.append(
                TeamMember(
                    name=name,
                    billing_line=billing_line,
                    category=category,
                    team=team,
                )
            )
        return members

    def _sprint_columns(self, context: _SheetContext, sprint: SprintWindow) -> list[int]:
        columns: list[int] = []
        for column_index, day in context.day_columns.items():
            day_date = date(context.window.year, context.window.month, day)
            if sprint.start_date <= day_date <= sprint.end_date:
                columns.append(column_index)
        return columns

    def _sprint_hours_between(
        self,
        start_date: date,
        end_date: date,
        holidays: set[date],
    ) -> float:
        total_hours = 0.0
        cursor = start_date
        while cursor <= end_date:
            if cursor.weekday() < 5 and cursor not in holidays:
                total_hours += self._day_hours_for_date(cursor)
            cursor = cursor + timedelta(days=1)
        return total_hours

    def _day_hours_for_date(self, day: date) -> float:
        if (day.month == 6 and day.day >= 15) or day.month in {7, 8} or (day.month == 9 and day.day <= 14):
            return 7.5
        return 8.5

    def _working_days_between(
        self,
        start_date: date,
        end_date: date,
        holidays: set[date],
    ) -> int:
        working_days = 0
        cursor = start_date
        while cursor <= end_date:
            if cursor.weekday() < 5 and cursor not in holidays:
                working_days += 1
            cursor = date.fromordinal(cursor.toordinal() + 1)
        return working_days

    def _normalize_team_name(self, value: str) -> str:
        normalized = " ".join(value.split())
        return TEAM_ALIASES.get(normalized, normalized)

    def _technician_column(self, headers: dict[str, int], required: bool) -> int | None:
        column = headers.get("Técnico")
        if column is not None:
            return column
        if required:
            raise ValueError("Technician column 'Técnico' not found in forecast sheet")
        return None

    def _fill_color(self, fill: PatternFill) -> int | None:
        fg_color = fill.fgColor
        if fg_color is None:
            return None
        if fg_color.type == "rgb" and fg_color.rgb is not None:
            return int(fg_color.rgb[-6:], 16)
        if fg_color.type == "indexed" and fg_color.indexed is not None:
            return fg_color.indexed
        return None

    def _is_holiday_fill(self, fill: PatternFill) -> bool:
        fg_color = fill.fgColor
        return (
            fill.patternType == "solid"
            and fg_color.type == "rgb"
            and fg_color.rgb is not None
            and fg_color.rgb[-6:] in HOLIDAY_FILL_RGB
        )

    def _is_non_working_fill(self, fill: PatternFill) -> bool:
        fg_color = fill.fgColor
        color_value = self._fill_color(fill)
        if color_value in NON_WORKING_FILL_COLORS:
            return True
        return (
            fill.patternType == "solid"
            and fg_color.type == "theme"
            and fg_color.theme == NON_WORKING_THEME
            and abs(fg_color.tint - NON_WORKING_TINT) < 1e-9
        )

    def _string_value(self, value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _merge_segments(
        self,
        segments: list[_SprintSegment],
        year: int,
        month: int,
    ) -> list[SprintWindow]:
        grouped: dict[tuple[str, str], list[_SprintSegment]] = {}
        for segment in segments:
            grouped.setdefault((segment.team, segment.sprint_id), []).append(segment)

        target_sheet_name = self._sheet_name(year, month)
        merged: list[SprintWindow] = []
        for key in sorted(grouped):
            group = grouped[key]
            group.sort(key=lambda item: item.start_date)
            target_segment = next(
                (segment for segment in group if segment.source_sheet == target_sheet_name),
                None,
            )
            source_sheet = target_segment.source_sheet if target_segment is not None else group[-1].source_sheet
            is_hatched = target_segment.is_hatched if target_segment is not None else group[-1].is_hatched
            merged.append(
                SprintWindow(
                    team=key[0],
                    sprint_id=key[1],
                    start_date=min(segment.start_date for segment in group),
                    end_date=max(segment.end_date for segment in group),
                    source_sheet=source_sheet,
                    is_hatched=is_hatched,
                )
            )

        return merged
