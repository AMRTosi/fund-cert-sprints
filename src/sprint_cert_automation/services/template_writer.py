from __future__ import annotations

from copy import copy
from datetime import datetime, time
import os
from pathlib import Path
from shutil import copy2
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from sprint_cert_automation.domain.models import CertificateDraft
from sprint_cert_automation.utils.dates import month_name_es

CONFIG_SHEET = "Config"
HOLIDAY_TABLE = "TB_Festivos"
TEAM_TABLE = "TB_Equipo"
PROFILE_TABLE = "TB_Perfiles"

PRODUCT_LABELS = {
    "Bonificaciones": "Bonificaciones",
    "Subvenciones": "Subvenciones",
    "Fondos de Reserva MRR": "Fondos de Reserva",
    "Fondos de Reserva": "Fondos de Reserva",
    "AccentureTransversal": "Transversales",
    "Transversal": "Transversales",
    "Transversales": "Transversales",
}

CATEGORY_ALIASES = {
    "responsable servicio": "Responsable del Servicio",
    "product owner": "Product Owner Proxy",
}


class TemplateWriter:
    """Generate certificate files from template and draft payload."""

    def __init__(self, template_path: Path) -> None:
        self.template_path = template_path

    def write(self, draft: CertificateDraft, output_path: Path) -> None:
        if self._should_use_com():
            self._write_with_com(draft, output_path)
            return

        self._write_with_openpyxl(draft, output_path)

    def _should_use_com(self) -> bool:
        if os.environ.get("PYTEST_CURRENT_TEST"):
            return False
        return os.name == "nt"

    def _write_with_openpyxl(self, draft: CertificateDraft, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        copy2(self.template_path, output_path)

        workbook = load_workbook(output_path, keep_vba=True)
        try:
            worksheet = workbook[CONFIG_SHEET]
            self._write_config(worksheet, draft)
            self._write_holidays(worksheet, draft)
            valid_categories = self._valid_categories(workbook)
            self._write_workloads(worksheet, draft, valid_categories)
            workbook.save(output_path)
        finally:
            vba_archive = getattr(workbook, "vba_archive", None)
            if vba_archive is not None:
                vba_archive.close()
            workbook.close()

    def _write_with_com(self, draft: CertificateDraft, output_path: Path) -> None:
        try:
            import win32com.client  # type: ignore
        except ImportError as exc:
            raise RuntimeError("pywin32 is required for Excel COM automation") from exc

        output_path.parent.mkdir(parents=True, exist_ok=True)
        copy2(self.template_path, output_path)

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = None
        try:
            workbook = excel.Workbooks.Open(str(output_path.resolve()))
            worksheet = workbook.Worksheets(CONFIG_SHEET)
            self._write_config_com(worksheet, draft)
            self._write_holidays_com(worksheet, draft)
            valid_categories = self._valid_categories_com(workbook)
            self._write_workloads_com(worksheet, draft, valid_categories)
            workbook.Save()
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=True)
            excel.Quit()

    def _write_config_com(self, worksheet, draft: CertificateDraft) -> None:
        worksheet.Range("A2").Value = self._to_excel_date(draft.start_date)
        worksheet.Range("B2").Value = self._to_excel_date(draft.end_date)
        worksheet.Range("D2").Value = self._display_sprint_id(draft)
        worksheet.Range("G3").Value = self._product_label(draft)

    def _write_holidays_com(self, worksheet, draft: CertificateDraft) -> None:
        table = worksheet.ListObjects(HOLIDAY_TABLE)
        data_range = table.DataBodyRange
        if data_range is not None:
            capacity = data_range.Rows.Count
            if len(draft.holidays) > capacity:
                raise ValueError(
                    f"Template table '{HOLIDAY_TABLE}' has capacity {capacity}, "
                    f"but {len(draft.holidays)} holidays were provided"
                )
            data_range.ClearContents()
        elif draft.holidays:
            raise ValueError(f"Template table '{HOLIDAY_TABLE}' has no data rows to write holidays")

        start_row = table.HeaderRowRange.Row + 1
        start_col = table.HeaderRowRange.Column
        for offset, holiday in enumerate(draft.holidays):
            row = start_row + offset
            worksheet.Cells(row, start_col).Value = holiday.label
            worksheet.Cells(row, start_col + 1).Value = self._to_excel_date(holiday.holiday_date)

    def _write_workloads_com(self, worksheet, draft: CertificateDraft, valid_categories: list[str]) -> None:
        table = worksheet.ListObjects(TEAM_TABLE)
        unique_workloads = self._dedupe_workloads(draft)
        data_range = table.DataBodyRange
        if data_range is not None:
            capacity = data_range.Rows.Count
            if len(unique_workloads) > capacity:
                raise ValueError(
                    f"Template table '{TEAM_TABLE}' has capacity {capacity}, "
                    f"but {len(unique_workloads)} workloads were provided"
                )
            data_range.ClearContents()
        elif unique_workloads:
            raise ValueError(f"Template table '{TEAM_TABLE}' has no data rows to write workloads")

        start_row = table.HeaderRowRange.Row + 1
        start_col = table.HeaderRowRange.Column
        for offset, workload in enumerate(unique_workloads):
            row = start_row + offset
            worksheet.Cells(row, start_col).Value = workload.member.name
            worksheet.Cells(row, start_col + 1).Value = workload.member.billing_line
            worksheet.Cells(row, start_col + 2).Value = self._resolve_category(
                workload.member.category,
                valid_categories,
            )
            worksheet.Cells(row, start_col + 3).Value = workload.sprint_hours
            worksheet.Cells(row, start_col + 4).Value = workload.free_hours

    def _valid_categories_com(self, workbook) -> list[str]:
        for worksheet in workbook.Worksheets:
            try:
                table = worksheet.ListObjects(PROFILE_TABLE)
            except Exception:
                continue

            header_row = table.HeaderRowRange.Row
            first_col = table.HeaderRowRange.Column
            last_col = table.Range.Column + table.Range.Columns.Count - 1
            last_row = table.Range.Row + table.Range.Rows.Count - 1

            category_column = None
            for column in range(first_col, last_col + 1):
                header = worksheet.Cells(header_row, column).Value
                if isinstance(header, str) and header.strip().casefold() == "categoria":
                    category_column = column
                    break
            if category_column is None:
                continue

            categories: list[str] = []
            for row in range(header_row + 1, last_row + 1):
                value = worksheet.Cells(row, category_column).Value
                if value is None:
                    continue
                label = str(value).strip()
                if label and label not in categories:
                    categories.append(label)
            if categories:
                return categories

        raise ValueError("Could not resolve category dropdown options from template")

    def _to_excel_date(self, value) -> datetime:
        return datetime.combine(value, time.min)

    def _write_config(self, worksheet, draft: CertificateDraft) -> None:
        worksheet["A2"] = draft.start_date
        worksheet["A2"].number_format = "dd-mm-yyyy"
        worksheet["B2"] = draft.end_date
        worksheet["B2"].number_format = "dd-mm-yyyy"
        worksheet["D2"] = self._display_sprint_id(draft)
        worksheet["G3"] = self._product_label(draft)

    def _write_holidays(self, worksheet, draft: CertificateDraft) -> None:
        start_row, end_row = self._table_data_bounds(worksheet, HOLIDAY_TABLE, 1)
        self._clear_rows(worksheet, start_row, end_row, 1, 2)
        rows_needed = max(len(draft.holidays), 1)
        self._expand_table_if_needed(worksheet, HOLIDAY_TABLE, rows_needed)
        start_row, end_row = self._table_data_bounds(worksheet, HOLIDAY_TABLE, rows_needed)
        self._clear_rows(worksheet, start_row, end_row, 1, 2)

        for offset, holiday in enumerate(draft.holidays):
            row = start_row + offset
            worksheet.cell(row=row, column=1, value=holiday.label)
            day_cell = worksheet.cell(row=row, column=2, value=holiday.holiday_date)
            day_cell.number_format = "dd-mm-yyyy"

    def _write_workloads(self, worksheet, draft: CertificateDraft, valid_categories: list[str]) -> None:
        start_row, end_row = self._table_data_bounds(worksheet, TEAM_TABLE, 1)
        self._clear_rows(worksheet, start_row, end_row, 6, 10)
        unique_workloads = self._dedupe_workloads(draft)
        rows_needed = max(len(unique_workloads), 1)
        self._expand_table_if_needed(worksheet, TEAM_TABLE, rows_needed)
        start_row, end_row = self._table_data_bounds(worksheet, TEAM_TABLE, rows_needed)
        self._clear_rows(worksheet, start_row, end_row, 6, 10)

        for offset, workload in enumerate(unique_workloads):
            row = start_row + offset
            worksheet.cell(row=row, column=6, value=workload.member.name)
            worksheet.cell(row=row, column=7, value=workload.member.billing_line)
            worksheet.cell(
                row=row,
                column=8,
                value=self._resolve_category(workload.member.category, valid_categories),
            )
            worksheet.cell(row=row, column=9, value=workload.sprint_hours)
            worksheet.cell(row=row, column=10, value=workload.free_hours)

    def _dedupe_workloads(self, draft: CertificateDraft):
        unique_workloads = []
        seen_names: set[str] = set()
        for workload in draft.workloads:
            key = self._normalize_key(workload.member.name)
            if key in seen_names:
                continue
            seen_names.add(key)
            unique_workloads.append(workload)
        return unique_workloads

    def _valid_categories(self, workbook) -> list[str]:
        for worksheet in workbook.worksheets:
            if PROFILE_TABLE not in worksheet.tables:
                continue
            table = worksheet.tables[PROFILE_TABLE]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            category_column = None
            for column in range(min_col, max_col + 1):
                header = worksheet.cell(row=min_row, column=column).value
                if isinstance(header, str) and header.strip().casefold() == "categoria":
                    category_column = column
                    break
            if category_column is None:
                continue

            categories: list[str] = []
            for row in range(min_row + 1, max_row + 1):
                value = worksheet.cell(row=row, column=category_column).value
                if value is None:
                    continue
                label = str(value).strip()
                if label and label not in categories:
                    categories.append(label)
            if categories:
                return categories
        raise ValueError("Could not resolve category dropdown options from template")

    def _resolve_category(self, category: str, valid_categories: list[str]) -> str:
        alias_target = CATEGORY_ALIASES.get(self._normalize_key(category))
        if alias_target is not None and alias_target in valid_categories:
            return alias_target

        normalized_options = {self._normalize_key(option): option for option in valid_categories}
        option = normalized_options.get(self._normalize_key(category))
        if option is not None:
            return option

        category_tokens = self._meaningful_tokens(category)
        for candidate in valid_categories:
            candidate_tokens = self._meaningful_tokens(candidate)
            if category_tokens and (category_tokens <= candidate_tokens or candidate_tokens <= category_tokens):
                return candidate

        for candidate in valid_categories:
            normalized_candidate = self._normalize_key(candidate)
            normalized_category = self._normalize_key(category)
            if normalized_category in normalized_candidate or normalized_candidate in normalized_category:
                return candidate

        raise ValueError(
            f"Category '{category}' is not available in template dropdown options"
        )

    def _normalize_key(self, value: str) -> str:
        collapsed = " ".join(value.split())
        no_accents = "".join(
            char for char in unicodedata.normalize("NFD", collapsed) if unicodedata.category(char) != "Mn"
        )
        return no_accents.casefold()

    def _meaningful_tokens(self, value: str) -> set[str]:
        stopwords = {"de", "del", "la", "el", "y"}
        return {token for token in self._normalize_key(value).split() if token and token not in stopwords}

    def _display_sprint_id(self, draft: CertificateDraft) -> str:
        if draft.team == "AccentureTransversal":
            return month_name_es(draft.end_date.month)
        return draft.sprint_id

    def _product_label(self, draft: CertificateDraft) -> str:
        if draft.product_label in PRODUCT_LABELS:
            return PRODUCT_LABELS[draft.product_label]
        return PRODUCT_LABELS.get(draft.team, draft.product_label)

    def _table_data_bounds(self, worksheet, table_name: str, minimum_rows: int) -> tuple[int, int]:
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        data_start_row = min_row + 1
        data_end_row = max(max_row, data_start_row + minimum_rows - 1)
        return data_start_row, data_end_row

    def _expand_table_if_needed(self, worksheet, table_name: str, data_rows: int) -> None:
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        required_end_row = min_row + max(data_rows, 1)
        if required_end_row <= max_row:
            return
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{required_end_row}"
        self._copy_row_style(worksheet, max_row, max_row + 1, min_col, max_col)
        for row in range(max_row + 2, required_end_row + 1):
            self._copy_row_style(worksheet, max_row + 1, row, min_col, max_col)

    def _copy_row_style(self, worksheet, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
        for column in range(min_col, max_col + 1):
            source_cell = worksheet.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=target_row, column=column)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

    def _clear_rows(self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        for row in range(start_row, end_row + 1):
            for column in range(start_col, end_col + 1):
                worksheet.cell(row=row, column=column).value = None
