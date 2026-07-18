"""Sprint configuration logic for period sheet duplication.

Reads sprint layout from a previous period sheet, calculates sprint boundaries
for a new month based on team rules, and writes the merged cells with proper
naming and fill styles.

Team rules (working days = Mon-Fri):
  - Bonificaciones: 10 working days, starts Tuesday, ends Monday.
  - Subvenciones: 16 working days (no fixed start day).
  - Fondos de Reserva: two sprints per month: days 1-15 and 16-last.
  - Transversal: one sprint spanning entire month.
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field

from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from sprint_cert_automation.services.sheet_duplicator import (
    DAY_LETTER_ROW,
    DAY_NUMBER_ROW,
    FIRST_DAY_COL,
)

# Sprint row assignments (1-based)
SPRINT_ROW_BONIF = 1
SPRINT_ROW_SUBV = 2
SPRINT_ROW_FDR = 3
SPRINT_ROW_TRANSV = 4

# Sprint duration in working days
BONIF_WORKING_DAYS = 10
SUBV_WORKING_DAYS = 16

# Spanish month names (lowercase) for Transversal naming
_MONTH_NAMES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}

# ── Fill styles per team ──────────────────────────────────────────────────

# Bonificaciones: theme 3, tint ~0.9 (light blue-ish)
BONIF_SOLID_FILL = PatternFill(
    patternType="solid",
    fgColor=Color(theme=3, tint=0.8999908444471572),
)
BONIF_HATCHED_FILL = PatternFill(
    patternType="darkDown",
    bgColor=Color(theme=3, tint=0.8999603259376812),
)

# Subvenciones: theme 5, tint ~0.8 (orange-ish)
SUBV_SOLID_FILL = PatternFill(
    patternType="solid",
    fgColor=Color(theme=5, tint=0.7999816888943144),
)
SUBV_HATCHED_FILL = PatternFill(
    patternType="lightDown",
    bgColor=Color(theme=5, tint=0.7999511703848384),
)

# Fondos de Reserva: solid RGB green
FDR_SOLID_FILL = PatternFill(
    patternType="solid",
    fgColor=Color(rgb="FFC0F1C8"),
)

# Transversal: theme 8, tint ~0.8 (gray/gold)
TRANSV_SOLID_FILL = PatternFill(
    patternType="solid",
    fgColor=Color(theme=8, tint=0.7999816888943144),
)


@dataclass
class SprintSegment:
    """One sprint segment within a month."""

    name: str
    sprint_number: int | None
    start_day: int        # 1-based calendar day
    end_day: int          # 1-based calendar day
    is_hatched: bool
    team: str             # "bonificaciones", "subvenciones", "fdr", "transversal"


@dataclass
class TeamSprintInfo:
    """Sprint info extracted from a sheet for a single team."""

    segments: list[SprintSegment] = field(default_factory=list)

    @property
    def last_sprint_number(self) -> int | None:
        if not self.segments:
            return None
        return max(s.sprint_number for s in self.segments if s.sprint_number is not None)

    @property
    def has_hatched_tail(self) -> bool:
        return bool(self.segments) and self.segments[-1].is_hatched

    @property
    def hatched_tail(self) -> SprintSegment | None:
        if self.has_hatched_tail:
            return self.segments[-1]
        return None


# ── Working-day utilities ─────────────────────────────────────────────────

def _count_working_days(year: int, month: int, day_start: int, day_end: int) -> int:
    """Count Mon-Fri days in [day_start, day_end] within the given month."""
    num_days_in_month = calendar.monthrange(year, month)[1]
    count = 0
    for d in range(day_start, min(day_end, num_days_in_month) + 1):
        if calendar.weekday(year, month, d) < 5:
            count += 1
    return count


def _advance_working_days(year: int, month: int, start_day: int, working_days: int) -> int:
    """Return the calendar day reached after advancing *working_days* working days
    from *start_day* (inclusive).  The returned day is the last working day consumed.

    If the working days extend beyond the month, return num_days_in_month + 1.
    """
    num_days_in_month = calendar.monthrange(year, month)[1]
    consumed = 0
    d = start_day
    while d <= num_days_in_month:
        if calendar.weekday(year, month, d) < 5:
            consumed += 1
            if consumed == working_days:
                return d
        d += 1
    return num_days_in_month + 1  # overflows


def _next_working_day(year: int, month: int, from_day: int) -> int:
    """Return the first Mon-Fri day on or after *from_day*.
    Returns num_days_in_month + 1 if none left.
    """
    num_days = calendar.monthrange(year, month)[1]
    d = from_day
    while d <= num_days:
        if calendar.weekday(year, month, d) < 5:
            return d
        d += 1
    return num_days + 1


# ── Read sprints from a sheet ─────────────────────────────────────────────

def _is_hatched_fill(fill) -> bool:
    """Return True if the fill uses a hatched pattern (lightDown, darkDown, etc.)."""
    return fill is not None and fill.patternType in (
        "lightDown", "darkDown", "lightUp", "darkUp",
        "lightGrid", "darkGrid", "lightHorizontal", "darkHorizontal",
        "lightVertical", "darkVertical",
    )


def _extract_sprint_number(name: str) -> int | None:
    """Extract the sprint number from a name like 'Bonificaciones - SP262' or 'SP262'."""
    if name is None:
        return None
    # Look for SPnnn pattern
    upper = name.upper()
    idx = upper.find("SP")
    if idx == -1:
        return None
    digits = ""
    for ch in upper[idx + 2:]:
        if ch.isdigit():
            digits += ch
        else:
            break
    return int(digits) if digits else None


def _team_for_row(row: int) -> str:
    return {
        SPRINT_ROW_BONIF: "bonificaciones",
        SPRINT_ROW_SUBV: "subvenciones",
        SPRINT_ROW_FDR: "fdr",
        SPRINT_ROW_TRANSV: "transversal",
    }.get(row, "unknown")


def read_sprints_from_sheet(ws: Worksheet) -> dict[str, TeamSprintInfo]:
    """Parse sprint segments from rows 1-4 of the given sheet.

    Returns a dict keyed by team name.
    """
    # Determine day-1 column
    day1_col = None
    for col in range(10, FIRST_DAY_COL + 1):
        if ws.cell(row=DAY_NUMBER_ROW, column=col).value == 1:
            day1_col = col
            break
    if day1_col is None:
        day1_col = FIRST_DAY_COL  # fallback

    # Find last day column
    last_day_col = day1_col
    for col in range(day1_col, 44):
        v = ws.cell(row=DAY_NUMBER_ROW, column=col).value
        if isinstance(v, (int, float)):
            last_day_col = col

    result: dict[str, TeamSprintInfo] = {
        "bonificaciones": TeamSprintInfo(),
        "subvenciones": TeamSprintInfo(),
        "fdr": TeamSprintInfo(),
        "transversal": TeamSprintInfo(),
    }

    for row in (SPRINT_ROW_BONIF, SPRINT_ROW_SUBV, SPRINT_ROW_FDR, SPRINT_ROW_TRANSV):
        team = _team_for_row(row)
        segments: list[SprintSegment] = []

        # Collect merged ranges for this row within the day columns
        for mr in ws.merged_cells.ranges:
            if mr.min_row != row or mr.max_row != row:
                continue
            if mr.max_col < day1_col or mr.min_col > last_day_col:
                continue
            cell = ws.cell(row=row, column=mr.min_col)
            val = cell.value
            if val is None:
                continue
            fill = cell.fill
            start_day = max(1, mr.min_col - day1_col + 1)
            end_day = mr.max_col - day1_col + 1
            segments.append(SprintSegment(
                name=str(val),
                sprint_number=_extract_sprint_number(str(val)),
                start_day=start_day,
                end_day=end_day,
                is_hatched=_is_hatched_fill(fill),
                team=team,
            ))

        # Collect unmerged cells
        for col in range(day1_col, last_day_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            in_merge = any(cell.coordinate in mr for mr in ws.merged_cells.ranges)
            if in_merge:
                continue
            fill = cell.fill
            day = col - day1_col + 1
            segments.append(SprintSegment(
                name=str(cell.value),
                sprint_number=_extract_sprint_number(str(cell.value)),
                start_day=day,
                end_day=day,
                is_hatched=_is_hatched_fill(fill),
                team=team,
            ))

        segments.sort(key=lambda s: s.start_day)
        result[team].segments = segments

    return result


# ── Calculate new sprint segments ─────────────────────────────────────────

def calculate_bonif_sprints(
    year: int, month: int, num_days: int,
    prev_info: TeamSprintInfo | None,
    prev_year: int | None, prev_month: int | None,
) -> list[SprintSegment]:
    """Calculate Bonificaciones sprint segments for the target month.

    Rules: 10 working days per sprint, starts Tuesday, ends Monday.
    """
    segments: list[SprintSegment] = []
    last_sp = prev_info.last_sprint_number if prev_info else None

    # Determine carry-over from previous month
    carry_working_days = 0
    carry_sp_num = None
    if prev_info and prev_info.has_hatched_tail and prev_year and prev_month:
        tail = prev_info.hatched_tail
        assert tail is not None
        carry_sp_num = tail.sprint_number
        consumed = _count_working_days(prev_year, prev_month, tail.start_day, tail.end_day)
        carry_working_days = max(0, BONIF_WORKING_DAYS - consumed)

    current_day = 1

    # Place carry-over segment
    if carry_working_days > 0 and carry_sp_num is not None:
        end_day = _advance_working_days(year, month, current_day, carry_working_days)
        if end_day > num_days:
            end_day = num_days
        segments.append(SprintSegment(
            name=f"SP{carry_sp_num}",
            sprint_number=carry_sp_num,
            start_day=current_day,
            end_day=end_day,
            is_hatched=False,
            team="bonificaciones",
        ))
        current_day = end_day + 1
        next_sp = carry_sp_num + 1
    else:
        next_sp = (last_sp + 1) if last_sp is not None else 1

    # Generate new sprints
    while current_day <= num_days:
        working_start = _next_working_day(year, month, current_day)
        if working_start > num_days:
            break
        end_day = _advance_working_days(year, month, working_start, BONIF_WORKING_DAYS)
        is_hatched = end_day > num_days
        actual_end = min(end_day, num_days)

        name = f"Bonificaciones - SP{next_sp}"
        if is_hatched:
            name = f"SP{next_sp}" if current_day == 1 and carry_working_days == 0 else f"Bonificaciones - SP{next_sp}"
            # Hatched sprints that start within the month use short name only
            # if they're at the tail
            name = f"Bonificaciones - SP{next_sp}"

        segments.append(SprintSegment(
            name=name,
            sprint_number=next_sp,
            start_day=current_day,
            end_day=actual_end,
            is_hatched=is_hatched,
            team="bonificaciones",
        ))
        current_day = actual_end + 1
        next_sp += 1

    # Adjust naming: carry-over segments use short "SPXXX" name (no team prefix)
    if segments and carry_working_days > 0:
        segments[0].name = f"SP{segments[0].sprint_number}"

    # Hatched tail segments also use shorter names in some sheets
    if segments and segments[-1].is_hatched:
        sp_num = segments[-1].sprint_number
        # Check if it's a short segment (few days) - use short name
        consumed = _count_working_days(year, month, segments[-1].start_day, segments[-1].end_day)
        if consumed < BONIF_WORKING_DAYS // 2:
            segments[-1].name = f"SP{sp_num}"
        else:
            segments[-1].name = f"Bonificaciones - SP{sp_num}"

    return segments


def calculate_subv_sprints(
    year: int, month: int, num_days: int,
    prev_info: TeamSprintInfo | None,
    prev_year: int | None, prev_month: int | None,
) -> list[SprintSegment]:
    """Calculate Subvenciones sprint segments for the target month.

    Rules: 16 working days per sprint.
    """
    segments: list[SprintSegment] = []
    last_sp = prev_info.last_sprint_number if prev_info else None

    carry_working_days = 0
    carry_sp_num = None
    if prev_info and prev_info.has_hatched_tail and prev_year and prev_month:
        tail = prev_info.hatched_tail
        assert tail is not None
        carry_sp_num = tail.sprint_number
        consumed = _count_working_days(prev_year, prev_month, tail.start_day, tail.end_day)
        carry_working_days = max(0, SUBV_WORKING_DAYS - consumed)

    current_day = 1

    if carry_working_days > 0 and carry_sp_num is not None:
        end_day = _advance_working_days(year, month, current_day, carry_working_days)
        if end_day > num_days:
            end_day = num_days
        segments.append(SprintSegment(
            name=f"SP{carry_sp_num}",
            sprint_number=carry_sp_num,
            start_day=current_day,
            end_day=end_day,
            is_hatched=False,
            team="subvenciones",
        ))
        current_day = end_day + 1
        next_sp = carry_sp_num + 1
    else:
        next_sp = (last_sp + 1) if last_sp is not None else 1

    while current_day <= num_days:
        working_start = _next_working_day(year, month, current_day)
        if working_start > num_days:
            break
        end_day = _advance_working_days(year, month, working_start, SUBV_WORKING_DAYS)
        is_hatched = end_day > num_days
        actual_end = min(end_day, num_days)

        name = f"Subvenciones - SP{next_sp}"

        segments.append(SprintSegment(
            name=name,
            sprint_number=next_sp,
            start_day=current_day,
            end_day=actual_end,
            is_hatched=is_hatched,
            team="subvenciones",
        ))
        current_day = actual_end + 1
        next_sp += 1

    # Adjust carry-over name
    if segments and carry_working_days > 0:
        segments[0].name = f"SP{segments[0].sprint_number}"

    # Hatched tail with few days → short name
    if segments and segments[-1].is_hatched:
        sp_num = segments[-1].sprint_number
        consumed = _count_working_days(year, month, segments[-1].start_day, segments[-1].end_day)
        if consumed < SUBV_WORKING_DAYS // 2:
            segments[-1].name = f"SP{sp_num}"
        else:
            segments[-1].name = f"Subvenciones - SP{sp_num}"

    return segments


def calculate_fdr_sprints(
    year: int, month: int, num_days: int,
    prev_info: TeamSprintInfo | None,
) -> list[SprintSegment]:
    """Calculate Fondos de Reserva sprint segments.

    Rules: always two sprints: days 1-15 and 16-last.
    """
    last_sp = prev_info.last_sprint_number if prev_info else None
    next_sp = (last_sp + 1) if last_sp is not None else 1

    return [
        SprintSegment(
            name=f"Fondos de Reserva - SP{next_sp}",
            sprint_number=next_sp,
            start_day=1,
            end_day=min(15, num_days),
            is_hatched=False,
            team="fdr",
        ),
        SprintSegment(
            name=f"Fondos de Reserva - SP{next_sp + 1}",
            sprint_number=next_sp + 1,
            start_day=16,
            end_day=num_days,
            is_hatched=False,
            team="fdr",
        ),
    ]


def calculate_transversal_sprint(
    year: int, month: int, num_days: int,
) -> list[SprintSegment]:
    """Calculate Transversal sprint segment.

    Rules: one sprint spanning the entire month, named with month name.
    """
    month_name = _MONTH_NAMES_ES[month]
    return [
        SprintSegment(
            name=f"Transversal - {month_name}",
            sprint_number=None,
            start_day=1,
            end_day=num_days,
            is_hatched=False,
            team="transversal",
        ),
    ]


# ── Write sprints to sheet ────────────────────────────────────────────────

def _get_fill(team: str, is_hatched: bool) -> PatternFill:
    """Return the appropriate fill for a team's sprint segment."""
    fills = {
        "bonificaciones": (BONIF_SOLID_FILL, BONIF_HATCHED_FILL),
        "subvenciones": (SUBV_SOLID_FILL, SUBV_HATCHED_FILL),
        "fdr": (FDR_SOLID_FILL, FDR_SOLID_FILL),
        "transversal": (TRANSV_SOLID_FILL, TRANSV_SOLID_FILL),
    }
    solid, hatched = fills.get(team, (PatternFill(), PatternFill()))
    return hatched if is_hatched else solid


def _team_row(team: str) -> int:
    return {
        "bonificaciones": SPRINT_ROW_BONIF,
        "subvenciones": SPRINT_ROW_SUBV,
        "fdr": SPRINT_ROW_FDR,
        "transversal": SPRINT_ROW_TRANSV,
    }[team]


def _clear_sprint_row(ws: Worksheet, row: int, day1_col: int, num_days: int) -> None:
    """Remove all merged cells and clear values/fills in a sprint row."""
    last_col = day1_col + num_days - 1

    # Collect merges to remove (can't modify during iteration)
    to_unmerge = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == row and mr.max_row == row:
            if mr.min_col >= day1_col and mr.min_col <= last_col:
                to_unmerge.append(str(mr))
    for mr_str in to_unmerge:
        ws.unmerge_cells(mr_str)

    # Clear cell values and fills
    for col in range(day1_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)


def write_sprints_to_sheet(
    ws: Worksheet,
    segments_by_team: dict[str, list[SprintSegment]],
    num_days: int,
) -> None:
    """Write sprint segments to rows 1-4, merging cells and applying fills."""
    day1_col = FIRST_DAY_COL

    for team, segments in segments_by_team.items():
        row = _team_row(team)
        _clear_sprint_row(ws, row, day1_col, num_days)

        for seg in segments:
            start_col = day1_col + seg.start_day - 1
            end_col = day1_col + seg.end_day - 1
            fill = _get_fill(team, seg.is_hatched)

            # Write name in the first cell
            cell = ws.cell(row=row, column=start_col)
            cell.value = seg.name
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

            # Apply fill to all cells in the range
            for col in range(start_col + 1, end_col + 1):
                c = ws.cell(row=row, column=col)
                c.fill = fill

            # Merge cells if the segment spans more than one column
            if end_col > start_col:
                ws.merge_cells(
                    start_row=row, start_column=start_col,
                    end_row=row, end_column=end_col,
                )


def configure_sprints(
    ws: Worksheet,
    year: int,
    month: int,
    num_days: int,
    prev_ws: Worksheet | None = None,
    prev_year: int | None = None,
    prev_month: int | None = None,
) -> dict[str, list[SprintSegment]]:
    """Calculate and write sprint segments for all four teams.

    If *prev_ws* is provided, reads sprint info from it to determine
    carry-over sprints and numbering.
    """
    prev_sprints = read_sprints_from_sheet(prev_ws) if prev_ws is not None else None

    segments_by_team: dict[str, list[SprintSegment]] = {}

    segments_by_team["bonificaciones"] = calculate_bonif_sprints(
        year, month, num_days,
        prev_info=prev_sprints["bonificaciones"] if prev_sprints else None,
        prev_year=prev_year, prev_month=prev_month,
    )
    segments_by_team["subvenciones"] = calculate_subv_sprints(
        year, month, num_days,
        prev_info=prev_sprints["subvenciones"] if prev_sprints else None,
        prev_year=prev_year, prev_month=prev_month,
    )
    segments_by_team["fdr"] = calculate_fdr_sprints(
        year, month, num_days,
        prev_info=prev_sprints["fdr"] if prev_sprints else None,
    )
    segments_by_team["transversal"] = calculate_transversal_sprint(
        year, month, num_days,
    )

    write_sprints_to_sheet(ws, segments_by_team, num_days)

    return segments_by_team
