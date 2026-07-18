"""Service for duplicating period sheets in the Forecast workbook."""

from __future__ import annotations

import calendar
import copy
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Calendar layout constants
DAY_NUMBER_ROW = 5
DAY_LETTER_ROW = 6
FIRST_DAY_COL = 13  # Column M
MAX_DAY_COL = 43  # Column AQ (day 31)
COST_FIRST_ROW = 7
COST_LAST_ROW = 37
REVENUE_FIRST_ROW = 41
REVENUE_LAST_ROW = 71

# Gray fill detection: theme=2, tint≈-0.1, solid pattern
_GRAY_THEME = 2
_GRAY_TINT_THRESHOLD = -0.05

# Spanish day-of-week letters (Monday=0 .. Sunday=6)
_WEEKDAY_LETTERS = ["L", "M", "X", "J", "V", "S", "D"]

NO_FILL = PatternFill(fill_type=None)

# Team column detection: the column that contains team names
_TEAM_VALUES = {"Transversal", "Bonificaciones", "Subvenciones", "Fondos de Reserva"}


@dataclass
class DuplicateSheetResult:
    """Result of a sheet duplication operation."""

    workbook_path: Path
    source_sheet: str
    new_sheet: str
    year: int
    month: int
    days_in_month: int


def _is_gray_fill(cell: Cell) -> bool:
    """Return True if the cell has a solid gray fill (theme 2, negative tint)."""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    return (
        fg.type == "theme"
        and fg.theme == _GRAY_THEME
        and fg.tint is not None
        and fg.tint < _GRAY_TINT_THRESHOLD
    )


def _day_col(day: int) -> int:
    """Return the 1-based column index for a given day number (1-31)."""
    return FIRST_DAY_COL + day - 1


def _generate_calendar(year: int, month: int) -> list[tuple[int, str]]:
    """Return list of (day_number, weekday_letter) for every day of the month."""
    num_days = calendar.monthrange(year, month)[1]
    result = []
    for day in range(1, num_days + 1):
        weekday = calendar.weekday(year, month, day)  # 0=Monday
        result.append((day, _WEEKDAY_LETTERS[weekday]))
    return result


def _get_cost_formula(cell: Cell, col_letter: str) -> str | None:
    """Extract the cost formula from a cell with a known formula, adjusting column.

    Returns None if the cell doesn't contain a recognized cost formula.
    """
    value = cell.value
    if not isinstance(value, str) or not value.startswith("="):
        return None
    return value


def _build_cost_formula_from_neighbor(
    ws: Worksheet, row: int, day_col: int, num_days: int
) -> str | None:
    """Find a neighboring cell in the same row that has a cost formula and adapt it.

    Scans the row for a cell with a formula to use as template.
    """
    for search_col in range(FIRST_DAY_COL, FIRST_DAY_COL + num_days):
        candidate = ws.cell(row=row, column=search_col)
        val = candidate.value
        if isinstance(val, str) and val.startswith("="):
            # Replace the source column letter with the target column letter
            src_letter = get_column_letter(search_col)
            tgt_letter = get_column_letter(day_col)
            # The formulas use patterns like IF(OR(M$6="S",...
            # Replace occurrences of the source column letter that are part of
            # cell references (not inside function names)
            adapted = _replace_column_in_formula(val, src_letter, tgt_letter)
            return adapted
    return None


def _replace_column_in_formula(formula: str, src_col: str, tgt_col: str) -> str:
    """Replace column references in a formula from src_col to tgt_col.

    Handles references like M$6, M7, $AT$3 (leaves $-prefixed columns alone).
    Only replaces the column letter when it appears as a cell reference prefix.
    """
    if src_col == tgt_col:
        return formula

    result = []
    i = 0
    while i < len(formula):
        # Check if we're at a position that matches the source column letter(s)
        if formula[i:i + len(src_col)] == src_col:
            # Verify it looks like a cell reference:
            # preceded by non-alpha (or start) and followed by $ or digit
            before_ok = i == 0 or not formula[i - 1].isalpha()
            after_pos = i + len(src_col)
            after_ok = after_pos < len(formula) and (
                formula[after_pos] == "$" or formula[after_pos].isdigit()
            )
            # Make sure we're not inside a longer column name (e.g., "AM" vs "A")
            # by checking the character before isn't a letter
            if before_ok and after_ok:
                result.append(tgt_col)
                i += len(src_col)
                continue
        result.append(formula[i])
        i += 1
    return "".join(result)


def update_calendar(ws: Worksheet, year: int, month: int) -> int:
    """Update day numbers (row 5) and day letters (row 6) for the given month.

    Returns the number of days in the month.
    """
    cal = _generate_calendar(year, month)
    num_days = len(cal)

    # Write day numbers and letters for each day of the month
    for day_num, day_letter in cal:
        col = _day_col(day_num)
        ws.cell(row=DAY_NUMBER_ROW, column=col, value=day_num)
        ws.cell(row=DAY_LETTER_ROW, column=col, value=day_letter)

    # Clear columns beyond the month's last day (e.g., days 29-31 for February)
    for day in range(num_days + 1, 32):
        col = _day_col(day)
        cell_num = ws.cell(row=DAY_NUMBER_ROW, column=col)
        cell_num.value = None
        cell_let = ws.cell(row=DAY_LETTER_ROW, column=col)
        cell_let.value = None
        # Clear all data rows for excess days
        for row in range(COST_FIRST_ROW, COST_LAST_ROW + 1):
            c = ws.cell(row=row, column=col)
            c.value = None
            c.fill = NO_FILL
        for row in range(REVENUE_FIRST_ROW, REVENUE_LAST_ROW + 1):
            c = ws.cell(row=row, column=col)
            c.value = None
            c.fill = NO_FILL

    return num_days


def remove_gray_fills(ws: Worksheet, num_days: int) -> int:
    """Remove gray background fills from T_COST_HOURS_ONLY (rows 7-37).

    Returns the number of cells cleared.
    """
    cleared = 0
    for row in range(COST_FIRST_ROW, COST_LAST_ROW + 1):
        for col in range(FIRST_DAY_COL, FIRST_DAY_COL + num_days):
            cell = ws.cell(row=row, column=col)
            if _is_gray_fill(cell):
                cell.fill = NO_FILL
                cleared += 1
    return cleared


def fill_empty_cost_cells(ws: Worksheet, num_days: int) -> int:
    """Fill empty cells in T_COST_HOURS_ONLY for working days with the row's formula.

    Only fills cells where the day letter (row 6) is a weekday (L, M, X, J, V).
    Returns the number of cells filled.
    """
    filled = 0
    for row in range(COST_FIRST_ROW, COST_LAST_ROW + 1):
        for col in range(FIRST_DAY_COL, FIRST_DAY_COL + num_days):
            cell = ws.cell(row=row, column=col)
            day_letter = ws.cell(row=DAY_LETTER_ROW, column=col).value

            # Skip weekends and cells that already have values
            if day_letter in ("S", "D"):
                continue
            if cell.value is not None:
                continue

            # Find a formula from a neighboring cell in the same row
            formula = _build_cost_formula_from_neighbor(ws, row, col, num_days)
            if formula is not None:
                cell.value = formula
                filled += 1

    return filled


def _find_team_column(ws: Worksheet) -> str:
    """Detect which column (letter) holds the team name (Equipo) in the sheet.

    Scans row-6 header for 'Equipo' first; if not found, checks rows 7-37
    for known team values in columns D, E, F.
    """
    # Check header row for 'Equipo'
    for col in range(1, 12):
        header = ws.cell(row=DAY_LETTER_ROW, column=col).value
        if header and "Equipo" in str(header):
            return get_column_letter(col)

    # Fallback: scan data rows for known team values
    for col in range(4, 7):  # D, E, F
        val = ws.cell(row=COST_FIRST_ROW, column=col).value
        if val and str(val) in _TEAM_VALUES:
            return get_column_letter(col)

    return "F"  # default


def _find_gap_anterior_columns(ws: Worksheet) -> list[int]:
    """Find all 'Gap Mes Anterior' column indices in row 6."""
    cols = []
    for col in range(44, 80):
        if ws.cell(row=DAY_LETTER_ROW, column=col).value == "Gap Mes Anterior":
            cols.append(col)
    return cols


def _find_prev_hatched_col_range(
    prev_ws: Worksheet, team: str,
) -> tuple[str, str] | None:
    """Find the column range of the hatched tail sprint for a team in the previous sheet.

    Returns (start_col_letter, end_col_letter) or None if no hatched tail exists.
    """
    from sprint_cert_automation.services.sprint_configurator import (
        read_sprints_from_sheet,
    )

    prev_sprints = read_sprints_from_sheet(prev_ws)
    team_info = prev_sprints.get(team)
    if not team_info or not team_info.has_hatched_tail:
        return None

    tail = team_info.hatched_tail
    assert tail is not None

    # Find day-1 column in previous sheet
    prev_day1_col = FIRST_DAY_COL
    for col in range(10, FIRST_DAY_COL + 1):
        if prev_ws.cell(row=DAY_NUMBER_ROW, column=col).value == 1:
            prev_day1_col = col
            break

    start_col = prev_day1_col + tail.start_day - 1
    end_col = prev_day1_col + tail.end_day - 1

    return (get_column_letter(start_col), get_column_letter(end_col))


def update_gap_anterior_formulas(
    ws: Worksheet,
    prev_sheet_name: str,
    prev_ws: Worksheet,
) -> int:
    """Update T_GAP_PERIODO_ANTERIOR formulas to reference the previous sheet.

    For each employee row (7-37), generates a formula that:
    - Returns 0 for Transversal and Fondos de Reserva teams.
    - Returns SUM of the corresponding T_REVENUES_TIMESHEET row (row+34) in
      the previous sheet's hatched tail range for Bonificaciones/Subvenciones.

    The revenue rows are used because the gap represents billable hours,
    not cost hours.

    Returns the number of cells updated.
    """
    # Revenue row offset: cost row 7 → revenue row 41, etc.
    REVENUE_OFFSET = REVENUE_FIRST_ROW - COST_FIRST_ROW  # 34

    team_col = _find_team_column(ws)
    gap_cols = _find_gap_anterior_columns(ws)

    if not gap_cols:
        return 0

    # The first Gap Mes Anterior column is the hours/revenue one
    gap_horas_col = gap_cols[0]

    # Find hatched tail ranges from previous sheet
    bonif_range = _find_prev_hatched_col_range(prev_ws, "bonificaciones")
    subv_range = _find_prev_hatched_col_range(prev_ws, "subvenciones")

    updated = 0
    for row in range(COST_FIRST_ROW, COST_LAST_ROW + 1):
        rev_row = row + REVENUE_OFFSET

        # Build the Bonificaciones branch (referencing revenue rows)
        if bonif_range:
            start_l, end_l = bonif_range
            if start_l == end_l:
                bonif_expr = f"'{prev_sheet_name}'!{start_l}{rev_row}"
            else:
                bonif_expr = f"SUM('{prev_sheet_name}'!{start_l}{rev_row}:{end_l}{rev_row})"
        else:
            bonif_expr = "0"

        # Build the Subvenciones branch (referencing revenue rows)
        if subv_range:
            start_l, end_l = subv_range
            if start_l == end_l:
                subv_expr = f"'{prev_sheet_name}'!{start_l}{rev_row}"
            else:
                subv_expr = f"SUM('{prev_sheet_name}'!{start_l}{rev_row}:{end_l}{rev_row})"
        else:
            subv_expr = "0"

        formula = (
            f'=IF({team_col}{row}="Transversal",0,'
            f'IF({team_col}{row}="Bonificaciones",{bonif_expr},'
            f'IF({team_col}{row}="Subvenciones",{subv_expr},'
            f'IF({team_col}{row}="Fondos de Reserva",0,0))))'
        )

        ws.cell(row=row, column=gap_horas_col).value = formula
        updated += 1

    # If there's a second Gap Mes Anterior (Factura), update it too
    # Its formula is: ={gap_horas_col_letter}{row}*{tarifa_col_letter}{row}
    if len(gap_cols) >= 2:
        gap_factura_col = gap_cols[1]
        gap_horas_letter = get_column_letter(gap_horas_col)

        # Find the Tarifa column (typically labeled "Tarifa" in row 6)
        tarifa_col_letter = None
        for col in range(1, 12):
            header = ws.cell(row=DAY_LETTER_ROW, column=col).value
            if header and "Tarifa" in str(header):
                tarifa_col_letter = get_column_letter(col)
                break

        if tarifa_col_letter is None:
            # Fallback: look at existing formula to determine tarifa column
            existing = ws.cell(row=COST_FIRST_ROW, column=gap_factura_col).value
            if existing and isinstance(existing, str) and "*" in existing:
                # Parse e.g. "=AW7*E7" to extract "E"
                parts = existing.replace("=", "").split("*")
                if len(parts) == 2:
                    ref = parts[1].strip()
                    tarifa_col_letter = "".join(c for c in ref if c.isalpha())

        if tarifa_col_letter:
            for row in range(COST_FIRST_ROW, COST_LAST_ROW + 1):
                formula = f"={gap_horas_letter}{row}*{tarifa_col_letter}{row}"
                ws.cell(row=row, column=gap_factura_col).value = formula
                updated += 1

    return updated


def duplicate_sheet(
    workbook_path: Path,
    source_sheet_name: str,
    new_sheet_name: str,
    year: int,
    month: int,
    previous_sheet_name: str | None = None,
    dry_run: bool = False,
) -> DuplicateSheetResult:
    """Duplicate a sheet in the forecast workbook and adapt it for the given month.

    Steps:
    1. Copy the source sheet and rename it.
    2. Position it right after the source sheet.
    3. Update the calendar (day numbers and weekday letters).
    4. Remove gray fills from T_COST_HOURS_ONLY.
    5. Fill empty cost cells on working days with the appropriate formula.
    6. Configure sprint segments (T_SPRINTS rows 1-4) based on previous period.
    7. Update T_GAP_PERIODO_ANTERIOR formulas to reference previous sheet.
    """
    from sprint_cert_automation.services.sprint_configurator import configure_sprints

    num_days = calendar.monthrange(year, month)[1]

    if dry_run:
        return DuplicateSheetResult(
            workbook_path=workbook_path,
            source_sheet=source_sheet_name,
            new_sheet=new_sheet_name,
            year=year,
            month=month,
            days_in_month=num_days,
        )

    wb = openpyxl.load_workbook(str(workbook_path))

    if source_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Source sheet '{source_sheet_name}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    if new_sheet_name in wb.sheetnames:
        raise ValueError(
            f"Sheet '{new_sheet_name}' already exists in the workbook."
        )

    if previous_sheet_name and previous_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Previous period sheet '{previous_sheet_name}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    # 1. Copy sheet
    source_ws = wb[source_sheet_name]
    new_ws = wb.copy_worksheet(source_ws)
    new_ws.title = new_sheet_name

    # 2. Position right after source
    source_idx = wb.sheetnames.index(source_sheet_name)
    wb.move_sheet(new_sheet_name, offset=source_idx + 1 - wb.sheetnames.index(new_sheet_name))

    # 3. Update calendar
    actual_days = update_calendar(new_ws, year, month)

    # 4. Remove gray fills
    remove_gray_fills(new_ws, actual_days)

    # 5. Fill empty cost cells
    fill_empty_cost_cells(new_ws, actual_days)

    # 6. Configure sprints
    prev_ws = None
    prev_year = None
    prev_month = None
    if previous_sheet_name:
        prev_ws = wb[previous_sheet_name]
        prev_year, prev_month = _infer_year_month_from_sheet_name(previous_sheet_name, year, month)

    configure_sprints(
        ws=new_ws,
        year=year,
        month=month,
        num_days=actual_days,
        prev_ws=prev_ws,
        prev_year=prev_year,
        prev_month=prev_month,
    )

    # 7. Update Gap Mes Anterior formulas
    if previous_sheet_name and prev_ws:
        update_gap_anterior_formulas(new_ws, previous_sheet_name, prev_ws)

    # Save
    wb.save(str(workbook_path))

    return DuplicateSheetResult(
        workbook_path=workbook_path,
        source_sheet=source_sheet_name,
        new_sheet=new_sheet_name,
        year=year,
        month=month,
        days_in_month=actual_days,
    )


def _infer_year_month_from_sheet_name(
    sheet_name: str, target_year: int, target_month: int
) -> tuple[int, int]:
    """Infer the calendar year and month from a FY sheet name.

    FY naming convention: FY{fiscal_year}_{month_abbr}
    Fiscal year starts September, e.g. FY27_sept = September 2026.
    Falls back to the month before the target if parsing fails.
    """
    month_map = {
        "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12,
        "ene": 1, "feb": 2, "mar": 3, "abr": 4,
        "mayo": 5, "may": 5, "jun": 6, "jul": 7, "ago": 8,
    }

    parts = sheet_name.split("_", 1)
    if len(parts) == 2:
        fy_part, month_part = parts
        fy_num_str = fy_part.replace("FY", "").replace("fy", "")
        month_key = month_part.lower().strip()

        if fy_num_str.isdigit() and month_key in month_map:
            fy_num = int(fy_num_str)
            m = month_map[month_key]
            # FY starts in September: Sept-Dec belong to (2000+fy_num-1),
            # Jan-Aug belong to (2000+fy_num)
            if m >= 9:
                y = 2000 + fy_num - 1
            else:
                y = 2000 + fy_num
            return y, m

    # Fallback: previous month of target
    if target_month == 1:
        return target_year - 1, 12
    return target_year, target_month - 1
