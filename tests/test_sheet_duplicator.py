"""Tests for the sheet duplicator service."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

from sprint_cert_automation.services.sheet_duplicator import (
    COST_FIRST_ROW,
    COST_LAST_ROW,
    DAY_LETTER_ROW,
    DAY_NUMBER_ROW,
    FIRST_DAY_COL,
    MAX_DAY_COL,
    REVENUE_FIRST_ROW,
    REVENUE_LAST_ROW,
    _find_gap_anterior_columns,
    _find_prev_hatched_col_range,
    _find_team_column,
    _generate_calendar,
    _is_gray_fill,
    _replace_column_in_formula,
    duplicate_sheet,
    fill_empty_cost_cells,
    remove_gray_fills,
    update_calendar,
    update_gap_anterior_formulas,
    update_revenues_mes_actual_formulas,
    update_revenues_no_fact_formulas,
    _find_gap_mes_actual_columns,
    _find_mes_actual_columns,
    _find_last_billable_day,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_template_workbook(path: Path, sheet_name: str = "Template_Mes") -> Path:
    """Create a minimal workbook that mirrors the Template_Mes structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Row 5: day numbers 1-31
    for day in range(1, 32):
        ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL + day - 1, value=day)

    # Row 6: day letters for a generic month (October-like: starts Thursday)
    letters = ["J", "V", "S", "D", "L", "M", "X", "J", "V", "S", "D",
               "L", "M", "X", "J", "V", "S", "D", "L", "M", "X",
               "J", "V", "S", "D", "L", "M", "X", "J", "V", "S"]
    for i, letter in enumerate(letters):
        ws.cell(row=DAY_LETTER_ROW, column=FIRST_DAY_COL + i, value=letter)

    # Header cells (row 6 cols A-K)
    ws.cell(row=6, column=1, value="Target")
    ws.cell(row=6, column=2, value="Nombre y Apellidos")

    # AT3 = 9 (hours config)
    ws.cell(row=3, column=46, value=9)

    # Rows 7-9: sample cost formulas (3 employees)
    for row in range(COST_FIRST_ROW, COST_FIRST_ROW + 3):
        ws.cell(row=row, column=1, value=1)  # Target
        ws.cell(row=row, column=2, value=f"Employee {row - 6}")
        for day in range(1, 32):
            col = FIRST_DAY_COL + day - 1
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.cell(
                row=row, column=col,
                value=f'=IF(OR({col_letter}$6="S", {col_letter}$6="D"), "",IF({col_letter}$6="V", 6.5, $AT$3))',
            )

    # Rows 41-43: sample revenue formulas
    for offset in range(3):
        cost_row = COST_FIRST_ROW + offset
        rev_row = REVENUE_FIRST_ROW + offset
        ws.cell(row=rev_row, column=2, value=f"Employee {offset + 1}")
        for day in range(1, 32):
            col = FIRST_DAY_COL + day - 1
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.cell(
                row=rev_row, column=col,
                value=f'=IF({col_letter}{cost_row}=9,8.5,IF({col_letter}{cost_row}=7,7.5,IF({col_letter}{cost_row}=6.5,IF($AT$3=9,8.5,7.5),"")))',
            )

    wb.save(str(path))
    return path


def _add_gray_cells(path: Path, sheet_name: str, cells: list[tuple[int, int]]) -> None:
    """Add gray fill to specific cells and clear their value."""
    wb = openpyxl.load_workbook(str(path))
    ws = wb[sheet_name]
    gray_fill = PatternFill(
        patternType="solid",
        fgColor=Color(theme=2, tint=-0.0999786370433668),
    )
    for row, col in cells:
        cell = ws.cell(row=row, column=col)
        cell.fill = gray_fill
        cell.value = None
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Calendar generation tests
# ---------------------------------------------------------------------------

class TestGenerateCalendar:
    def test_february_2026_has_28_days(self):
        cal = _generate_calendar(2026, 2)
        assert len(cal) == 28
        # Feb 1 2026 is a Sunday
        assert cal[0] == (1, "D")
        assert cal[1] == (2, "L")
        assert cal[27] == (28, "S")

    def test_february_2028_leap_year_has_29_days(self):
        cal = _generate_calendar(2028, 2)
        assert len(cal) == 29

    def test_october_2026_has_31_days(self):
        cal = _generate_calendar(2026, 10)
        assert len(cal) == 31
        # Oct 1 2026 is a Thursday
        assert cal[0] == (1, "J")

    def test_november_2026_has_30_days(self):
        cal = _generate_calendar(2026, 11)
        assert len(cal) == 30
        # Nov 1 2026 is a Sunday
        assert cal[0] == (1, "D")

    def test_weekday_letters_are_spanish(self):
        cal = _generate_calendar(2026, 10)
        # Oct 5 2026 is Monday
        assert cal[4] == (5, "L")
        # Oct 7 2026 is Wednesday
        assert cal[6] == (7, "X")


# ---------------------------------------------------------------------------
# Calendar update tests
# ---------------------------------------------------------------------------

class TestUpdateCalendar:
    def test_updates_day_numbers_and_letters(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        num_days = update_calendar(ws, 2026, 11)  # November 2026

        assert num_days == 30

        # Check day 1
        assert ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL).value == 1
        # Nov 1 2026 is Sunday
        assert ws.cell(row=DAY_LETTER_ROW, column=FIRST_DAY_COL).value == "D"

        # Check day 30
        col_30 = FIRST_DAY_COL + 29
        assert ws.cell(row=DAY_NUMBER_ROW, column=col_30).value == 30

        # Day 31 should be cleared
        col_31 = FIRST_DAY_COL + 30
        assert ws.cell(row=DAY_NUMBER_ROW, column=col_31).value is None
        assert ws.cell(row=DAY_LETTER_ROW, column=col_31).value is None

    def test_clears_data_for_excess_days_february(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        num_days = update_calendar(ws, 2026, 2)  # February 2026 = 28 days

        assert num_days == 28

        # Days 29, 30, 31 should be cleared
        for excess_day in [29, 30, 31]:
            col = FIRST_DAY_COL + excess_day - 1
            assert ws.cell(row=DAY_NUMBER_ROW, column=col).value is None
            # Cost rows should be cleared too
            assert ws.cell(row=COST_FIRST_ROW, column=col).value is None

    def test_31_day_month_keeps_all_columns(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        num_days = update_calendar(ws, 2026, 10)  # October 2026 = 31 days

        assert num_days == 31
        col_31 = FIRST_DAY_COL + 30
        assert ws.cell(row=DAY_NUMBER_ROW, column=col_31).value == 31


# ---------------------------------------------------------------------------
# Gray fill tests
# ---------------------------------------------------------------------------

class TestRemoveGrayFills:
    def test_removes_gray_fills_from_cost_area(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        gray_cells = [
            (COST_FIRST_ROW, FIRST_DAY_COL + 1),  # day 2
            (COST_FIRST_ROW, FIRST_DAY_COL + 4),  # day 5
            (COST_FIRST_ROW + 1, FIRST_DAY_COL + 8),  # day 9
        ]
        _add_gray_cells(wb_path, "Template_Mes", gray_cells)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        cleared = remove_gray_fills(ws, 31)

        assert cleared == 3
        for row, col in gray_cells:
            assert not _is_gray_fill(ws.cell(row=row, column=col))

    def test_does_not_touch_non_gray_cells(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        cleared = remove_gray_fills(ws, 31)

        assert cleared == 0


# ---------------------------------------------------------------------------
# Fill empty cost cells tests
# ---------------------------------------------------------------------------

class TestFillEmptyCostCells:
    def test_fills_empty_weekday_cells_with_formula(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        # Add gray cells (which clears value) then remove fills
        gray_cells = [
            (COST_FIRST_ROW, FIRST_DAY_COL + 4),  # day 5, should be weekday
        ]
        _add_gray_cells(wb_path, "Template_Mes", gray_cells)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        # Set day 5 to a weekday letter (Monday)
        ws.cell(row=DAY_LETTER_ROW, column=FIRST_DAY_COL + 4, value="L")

        remove_gray_fills(ws, 31)
        filled = fill_empty_cost_cells(ws, 31)

        assert filled == 1
        cell = ws.cell(row=COST_FIRST_ROW, column=FIRST_DAY_COL + 4)
        assert cell.value is not None
        assert cell.value.startswith("=")

    def test_does_not_fill_weekend_cells(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        # Manually clear a weekend cell
        ws.cell(row=DAY_LETTER_ROW, column=FIRST_DAY_COL + 2).value = "S"
        ws.cell(row=COST_FIRST_ROW, column=FIRST_DAY_COL + 2).value = None

        filled = fill_empty_cost_cells(ws, 31)

        # Should not fill cells on Saturday
        cell = ws.cell(row=COST_FIRST_ROW, column=FIRST_DAY_COL + 2)
        assert cell.value is None

    def test_does_not_overwrite_existing_formulas(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb.active

        original_value = ws.cell(row=COST_FIRST_ROW, column=FIRST_DAY_COL).value
        filled = fill_empty_cost_cells(ws, 31)

        assert ws.cell(row=COST_FIRST_ROW, column=FIRST_DAY_COL).value == original_value


# ---------------------------------------------------------------------------
# Formula column replacement tests
# ---------------------------------------------------------------------------

class TestReplaceColumnInFormula:
    def test_replaces_simple_column_reference(self):
        formula = '=IF(OR(M$6="S", M$6="D"), "",IF(M$6="V", 6.5, $AT$3))'
        result = _replace_column_in_formula(formula, "M", "N")
        assert result == '=IF(OR(N$6="S", N$6="D"), "",IF(N$6="V", 6.5, $AT$3))'

    def test_replaces_two_letter_column(self):
        formula = '=IF(OR(AO$6="S", AO$6="D"), "",IF(AO$6="V", 6.5, $AT$3))'
        result = _replace_column_in_formula(formula, "AO", "AP")
        assert result == '=IF(OR(AP$6="S", AP$6="D"), "",IF(AP$6="V", 6.5, $AT$3))'

    def test_does_not_replace_dollar_prefixed_refs(self):
        formula = '=IF(M$6="V", 6.5, $AT$3)'
        result = _replace_column_in_formula(formula, "M", "N")
        # $AT$3 should remain unchanged
        assert "$AT$3" in result

    def test_same_column_returns_unchanged(self):
        formula = '=IF(M$6="S", "", 8)'
        result = _replace_column_in_formula(formula, "M", "M")
        assert result == formula


# ---------------------------------------------------------------------------
# Gray fill detection tests
# ---------------------------------------------------------------------------

class TestIsGrayFill:
    def test_detects_gray_fill(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        cell.fill = PatternFill(
            patternType="solid",
            fgColor=Color(theme=2, tint=-0.0999786370433668),
        )
        assert _is_gray_fill(cell) is True

    def test_rejects_no_fill(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert _is_gray_fill(cell) is False

    def test_rejects_non_gray_solid_fill(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        cell.fill = PatternFill(patternType="solid", fgColor="FF0000")
        assert _is_gray_fill(cell) is False


# ---------------------------------------------------------------------------
# End-to-end duplicate_sheet tests
# ---------------------------------------------------------------------------

class TestDuplicateSheet:
    def test_dry_run_does_not_modify_file(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)
        original_size = wb_path.stat().st_size

        result = duplicate_sheet(
            workbook_path=wb_path,
            source_sheet_name="Template_Mes",
            new_sheet_name="FY27_dic",
            year=2026,
            month=12,
            dry_run=True,
        )

        assert result.new_sheet == "FY27_dic"
        assert result.days_in_month == 31
        # File should not have been modified
        wb = openpyxl.load_workbook(str(wb_path))
        assert "FY27_dic" not in wb.sheetnames

    def test_creates_new_sheet_with_correct_name(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        result = duplicate_sheet(
            workbook_path=wb_path,
            source_sheet_name="Template_Mes",
            new_sheet_name="FY27_dic",
            year=2026,
            month=12,
        )

        wb = openpyxl.load_workbook(str(wb_path))
        assert "FY27_dic" in wb.sheetnames
        assert result.new_sheet == "FY27_dic"

    def test_new_sheet_positioned_after_source(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        # Add another sheet after Template_Mes
        wb = openpyxl.load_workbook(str(wb_path))
        wb.create_sheet("ExtraSheet")
        wb.save(str(wb_path))

        duplicate_sheet(
            workbook_path=wb_path,
            source_sheet_name="Template_Mes",
            new_sheet_name="FY27_dic",
            year=2026,
            month=12,
        )

        wb = openpyxl.load_workbook(str(wb_path))
        names = wb.sheetnames
        src_idx = names.index("Template_Mes")
        new_idx = names.index("FY27_dic")
        assert new_idx == src_idx + 1

    def test_calendar_updated_for_target_month(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        duplicate_sheet(
            workbook_path=wb_path,
            source_sheet_name="Template_Mes",
            new_sheet_name="FY27_dic",
            year=2026,
            month=12,
        )

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb["FY27_dic"]

        # Dec 1 2026 is Tuesday
        assert ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL).value == 1
        assert ws.cell(row=DAY_LETTER_ROW, column=FIRST_DAY_COL).value == "M"

    def test_february_clears_excess_days(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        duplicate_sheet(
            workbook_path=wb_path,
            source_sheet_name="Template_Mes",
            new_sheet_name="FY26_feb",
            year=2026,
            month=2,
        )

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb["FY26_feb"]

        # Day 29 should be cleared
        col_29 = FIRST_DAY_COL + 28
        assert ws.cell(row=DAY_NUMBER_ROW, column=col_29).value is None
        assert ws.cell(row=COST_FIRST_ROW, column=col_29).value is None

    def test_gray_fills_removed_and_formulas_filled(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        # Add gray cells on weekdays
        gray_cells = [
            (COST_FIRST_ROW, FIRST_DAY_COL + 4),  # day 5
        ]
        _add_gray_cells(wb_path, "Template_Mes", gray_cells)

        duplicate_sheet(
            workbook_path=wb_path,
            source_sheet_name="Template_Mes",
            new_sheet_name="FY27_oct",
            year=2026,
            month=10,
        )

        wb = openpyxl.load_workbook(str(wb_path))
        ws = wb["FY27_oct"]

        # Oct 5 2026 is Monday (L) - should have formula filled
        cell = ws.cell(row=COST_FIRST_ROW, column=FIRST_DAY_COL + 4)
        day_letter = ws.cell(row=DAY_LETTER_ROW, column=FIRST_DAY_COL + 4).value
        if day_letter not in ("S", "D"):
            assert cell.value is not None
            assert str(cell.value).startswith("=")

    def test_raises_on_missing_source_sheet(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        with pytest.raises(ValueError, match="not found"):
            duplicate_sheet(
                workbook_path=wb_path,
                source_sheet_name="NonExistent",
                new_sheet_name="FY27_dic",
                year=2026,
                month=12,
            )

    def test_raises_on_existing_target_sheet(self, tmp_path):
        wb_path = tmp_path / "test.xlsx"
        _create_template_workbook(wb_path)

        with pytest.raises(ValueError, match="already exists"):
            duplicate_sheet(
                workbook_path=wb_path,
                source_sheet_name="Template_Mes",
                new_sheet_name="Template_Mes",
                year=2026,
                month=12,
            )


# ---------------------------------------------------------------------------
# CLI integration test for duplicate-sheet
# ---------------------------------------------------------------------------

class TestDuplicateSheetCli:
    def test_cli_duplicate_sheet_dry_run(self, monkeypatch, tmp_path):
        from sprint_cert_automation import cli

        captured = {}

        def fake_duplicate(**kwargs):
            captured.update(kwargs)

            class Result:
                workbook_path = kwargs["forecast_path"]
                source_sheet = kwargs["source_sheet"]
                new_sheet = kwargs["new_sheet"]
                year = kwargs["year"]
                month = kwargs["month"]
                days_in_month = 31

            return Result()

        monkeypatch.setattr(cli, "duplicate_period_sheet", fake_duplicate)

        exit_code = cli.main([
            "duplicate-sheet",
            "--forecast", "forecast.xlsx",
            "--source", "Template_Mes",
            "--target", "FY27_dic",
            "--year", "2026",
            "--month", "12",
            "--dry-run",
        ])

        assert exit_code == 0
        assert captured["source_sheet"] == "Template_Mes"
        assert captured["new_sheet"] == "FY27_dic"
        assert captured["year"] == 2026
        assert captured["month"] == 12
        assert captured["dry_run"] is True


# ---------------------------------------------------------------------------
# Gap Mes Anterior formula tests
# ---------------------------------------------------------------------------


def _create_prev_sheet_with_hatched_sprints(wb, name, year, month,
                                             bonif_hatched_range=None,
                                             subv_hatched_range=None):
    """Create a previous period sheet with optional hatched tail sprints."""
    from sprint_cert_automation.services.sprint_configurator import (
        BONIF_HATCHED_FILL,
        BONIF_SOLID_FILL,
        SUBV_HATCHED_FILL,
        SUBV_SOLID_FILL,
        FDR_SOLID_FILL,
        TRANSV_SOLID_FILL,
    )

    ws = wb.create_sheet(name)
    # Set up calendar row 5 (day numbers)
    import calendar as cal_mod
    num_days = cal_mod.monthrange(year, month)[1]
    for day in range(1, num_days + 1):
        ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL + day - 1, value=day)

    def _write_segment(row, start_day, end_day, label, fill):
        start_col = FIRST_DAY_COL + start_day - 1
        end_col = FIRST_DAY_COL + end_day - 1
        cell = ws.cell(row=row, column=start_col)
        cell.value = label
        cell.fill = fill
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=row, column=c).fill = fill
        if end_col > start_col:
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col)

    # Bonificaciones (row 1)
    if bonif_hatched_range:
        h_start, h_end = bonif_hatched_range
        if h_start > 1:
            _write_segment(1, 1, h_start - 1, "SP263", BONIF_SOLID_FILL)
        _write_segment(1, h_start, h_end, "SP264", BONIF_HATCHED_FILL)
    else:
        _write_segment(1, 1, num_days, "SP263", BONIF_SOLID_FILL)

    # Subvenciones (row 2)
    if subv_hatched_range:
        h_start, h_end = subv_hatched_range
        if h_start > 1:
            _write_segment(2, 1, h_start - 1, "SP206", SUBV_SOLID_FILL)
        _write_segment(2, h_start, h_end, "SP207", SUBV_HATCHED_FILL)
    else:
        _write_segment(2, 1, num_days, "SP206", SUBV_SOLID_FILL)

    # FdR and Transversal always complete in-month
    _write_segment(3, 1, num_days, "FdR Sprint 1", FDR_SOLID_FILL)
    _write_segment(4, 1, num_days, "TRANSVERSAL", TRANSV_SOLID_FILL)

    return ws


def _create_sheet_with_gap_cols(wb, name, gap_col, gap_factura_col=None):
    """Create a sheet with Gap Mes Anterior column(s) and team column."""
    ws = wb.create_sheet(name)

    # Set up calendar
    for day in range(1, 32):
        ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL + day - 1, value=day)

    # Team column at F (like Template_Mes)
    ws.cell(row=DAY_LETTER_ROW, column=6, value="Equipo")

    # Tarifa column at G
    ws.cell(row=DAY_LETTER_ROW, column=7, value="Tarifa")

    # Set team names for rows 7-9
    teams = ["Transversal", "Bonificaciones", "Subvenciones"]
    for i, team in enumerate(teams):
        ws.cell(row=COST_FIRST_ROW + i, column=6, value=team)
        ws.cell(row=COST_FIRST_ROW + i, column=7, value=50)  # tarifa rate

    # Gap Mes Anterior header
    ws.cell(row=DAY_LETTER_ROW, column=gap_col, value="Gap Mes Anterior")

    if gap_factura_col:
        ws.cell(row=DAY_LETTER_ROW, column=gap_factura_col, value="Gap Mes Anterior")

    return ws


class TestFindTeamColumn:
    def test_finds_equipo_header(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=6, value="Equipo")
        assert _find_team_column(ws) == "F"

    def test_finds_equipo_at_col_e(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=5, value="Equipo")
        assert _find_team_column(ws) == "E"

    def test_fallback_to_data_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=COST_FIRST_ROW, column=4, value="Transversal")
        assert _find_team_column(ws) == "D"

    def test_default_when_no_team(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _find_team_column(ws) == "F"


class TestFindGapAnteriorColumns:
    def test_finds_single_gap_col(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=52, value="Gap Mes Anterior")
        assert _find_gap_anterior_columns(ws) == [52]

    def test_finds_two_gap_cols(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=49, value="Gap Mes Anterior")
        ws.cell(row=DAY_LETTER_ROW, column=55, value="Gap Mes Anterior")
        assert _find_gap_anterior_columns(ws) == [49, 55]

    def test_no_gap_cols(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert _find_gap_anterior_columns(ws) == []


class TestFindPrevHatchedColRange:
    def test_bonif_hatched_tail(self):
        wb = openpyxl.Workbook()
        ws = _create_prev_sheet_with_hatched_sprints(
            wb, "prev", 2026, 10, bonif_hatched_range=(28, 31))
        result = _find_prev_hatched_col_range(ws, "bonificaciones")
        assert result is not None
        start_l, end_l = result
        # Day 28 → col 13+27=40 → AN, Day 31 → col 13+30=43 → AQ
        assert start_l == "AN"
        assert end_l == "AQ"

    def test_subv_hatched_tail(self):
        wb = openpyxl.Workbook()
        ws = _create_prev_sheet_with_hatched_sprints(
            wb, "prev", 2026, 10, subv_hatched_range=(29, 31))
        result = _find_prev_hatched_col_range(ws, "subvenciones")
        assert result is not None
        start_l, end_l = result
        # Day 29 → col 13+28=41 → AO, Day 31 → col 13+30=43 → AQ
        assert start_l == "AO"
        assert end_l == "AQ"

    def test_no_hatched_tail(self):
        wb = openpyxl.Workbook()
        ws = _create_prev_sheet_with_hatched_sprints(wb, "prev", 2026, 10)
        result = _find_prev_hatched_col_range(ws, "bonificaciones")
        assert result is None

    def test_fdr_never_hatched(self):
        wb = openpyxl.Workbook()
        ws = _create_prev_sheet_with_hatched_sprints(wb, "prev", 2026, 10)
        result = _find_prev_hatched_col_range(ws, "fdr")
        assert result is None

    def test_single_day_hatched(self):
        wb = openpyxl.Workbook()
        ws = _create_prev_sheet_with_hatched_sprints(
            wb, "prev", 2026, 10, bonif_hatched_range=(31, 31))
        result = _find_prev_hatched_col_range(ws, "bonificaciones")
        assert result is not None
        start_l, end_l = result
        assert start_l == end_l == "AQ"


class TestUpdateGapAnteriorFormulas:
    def test_basic_formula_generation(self):
        """Generates correct IF formulas with SUM ranges for hatched tails."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10,
            bonif_hatched_range=(28, 31),
            subv_hatched_range=(29, 31))
        new_ws = _create_sheet_with_gap_cols(wb, "FY26_nov", gap_col=49)

        updated = update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)
        assert updated == 31  # rows 7-37

        # Check row 7 (Transversal → 0)
        f7 = new_ws.cell(row=7, column=49).value
        assert '"Transversal",0' in f7

        # Check row 8 (Bonificaciones → SUM of hatched range in revenue rows)
        # Cost row 8 → revenue row 42 in previous sheet
        f8 = new_ws.cell(row=8, column=49).value
        assert "SUM('FY26_oct'!AN42:AQ42)" in f8

        # Check row 9 (Subvenciones → SUM of hatched range in revenue rows)
        # Cost row 9 → revenue row 43 in previous sheet
        f9 = new_ws.cell(row=9, column=49).value
        assert "SUM('FY26_oct'!AO43:AQ43)" in f9

    def test_formula_with_no_hatched_tails(self):
        """When no team has hatched tails, formulas use 0."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10)
        new_ws = _create_sheet_with_gap_cols(wb, "FY26_nov", gap_col=49)

        updated = update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)
        assert updated == 31

        f8 = new_ws.cell(row=8, column=49).value
        assert '"Bonificaciones",0' in f8
        assert '"Subvenciones",0' in f8

    def test_single_cell_reference(self):
        """Single-day hatched tail uses cell ref instead of SUM (revenue row)."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10,
            bonif_hatched_range=(31, 31))
        new_ws = _create_sheet_with_gap_cols(wb, "FY26_nov", gap_col=49)

        update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)

        # Cost row 8 → revenue row 42
        f8 = new_ws.cell(row=8, column=49).value
        assert "'FY26_oct'!AQ42" in f8
        assert "SUM" not in f8

    def test_fdr_always_zero(self):
        """FdR team always returns 0 regardless of previous sheet."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10,
            bonif_hatched_range=(28, 31))
        new_ws = _create_sheet_with_gap_cols(wb, "FY26_nov", gap_col=49)
        # Add FdR team row
        new_ws.cell(row=10, column=6, value="Fondos de Reserva")

        update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)

        f10 = new_ws.cell(row=10, column=49).value
        assert '"Fondos de Reserva",0' in f10

    def test_factura_column_formula(self):
        """Second Gap Mes Anterior column gets hours × tarifa formula."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10,
            bonif_hatched_range=(28, 31))
        new_ws = _create_sheet_with_gap_cols(
            wb, "FY26_nov", gap_col=49, gap_factura_col=55)

        updated = update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)
        # 31 horas + 31 factura = 62
        assert updated == 62

        # Factura formula: gap_horas_col_letter is AW (col 49)
        f7_factura = new_ws.cell(row=7, column=55).value
        assert f7_factura == "=AW7*G7"

    def test_no_gap_columns_returns_zero(self):
        """Returns 0 when no Gap Mes Anterior columns found."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10)
        new_ws = wb.create_sheet("FY26_nov")

        updated = update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)
        assert updated == 0

    def test_team_column_detection_in_formula(self):
        """Formula uses the correct team column letter."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10)
        new_ws = _create_sheet_with_gap_cols(wb, "FY26_nov", gap_col=49)
        # Team column is at F (from Equipo header)

        update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)

        f7 = new_ws.cell(row=7, column=49).value
        assert f7.startswith('=IF(F7="Transversal"')

    def test_formula_references_revenue_rows_not_cost(self):
        """Formula references revenue rows (row+34) in previous sheet."""
        wb = openpyxl.Workbook()
        prev_ws = _create_prev_sheet_with_hatched_sprints(
            wb, "FY26_oct", 2026, 10,
            bonif_hatched_range=(28, 31))
        new_ws = _create_sheet_with_gap_cols(wb, "FY26_nov", gap_col=49)

        update_gap_anterior_formulas(new_ws, "FY26_oct", prev_ws)

        # Row 7 (first cost row) → references revenue row 41
        f7 = new_ws.cell(row=7, column=49).value
        assert "AN41:AQ41" in f7 or "AN41" in f7

        # Row 37 (last cost row) → references revenue row 71
        f37 = new_ws.cell(row=37, column=49).value
        assert "AN71:AQ71" in f37 or "AN71" in f37


# ---------------------------------------------------------------------------
# Revenues Mes Actual formula tests
# ---------------------------------------------------------------------------


def _create_sheet_with_sprints_and_mes_actual(wb, name, year, month,
                                               bonif_segments=None,
                                               subv_segments=None,
                                               mes_actual_col=53,
                                               mes_actual_factura_col=None):
    """Create a sheet with sprint config and Mes Actual column for testing."""
    from sprint_cert_automation.services.sprint_configurator import (
        BONIF_HATCHED_FILL,
        BONIF_SOLID_FILL,
        SUBV_HATCHED_FILL,
        SUBV_SOLID_FILL,
        FDR_SOLID_FILL,
        TRANSV_SOLID_FILL,
    )
    import calendar as cal_mod

    ws = wb.create_sheet(name)
    num_days = cal_mod.monthrange(year, month)[1]

    # Set up calendar row 5 (day numbers)
    for day in range(1, num_days + 1):
        ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL + day - 1, value=day)

    # Team column at F
    ws.cell(row=DAY_LETTER_ROW, column=6, value="Equipo")
    # Tarifa column at G
    ws.cell(row=DAY_LETTER_ROW, column=7, value="Tarifa")

    # Set team names and tarifa for rows 7-10
    teams = ["Transversal", "Bonificaciones", "Subvenciones", "Fondos de Reserva"]
    for i, team in enumerate(teams):
        ws.cell(row=COST_FIRST_ROW + i, column=6, value=team)
        ws.cell(row=COST_FIRST_ROW + i, column=7, value=50)
        # Also set revenue row team
        ws.cell(row=REVENUE_FIRST_ROW + i, column=6, value=team)

    # Mes Actual header
    ws.cell(row=DAY_LETTER_ROW, column=mes_actual_col, value="Mes Actual")
    if mes_actual_factura_col:
        ws.cell(row=DAY_LETTER_ROW, column=mes_actual_factura_col, value="Mes Actual")

    def _write_segment(row, start_day, end_day, label, fill):
        start_col = FIRST_DAY_COL + start_day - 1
        end_col = FIRST_DAY_COL + end_day - 1
        cell = ws.cell(row=row, column=start_col)
        cell.value = label
        cell.fill = fill
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=row, column=c).fill = fill
        if end_col > start_col:
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col)

    # Bonificaciones (row 1)
    if bonif_segments:
        for start, end, label, hatched in bonif_segments:
            fill = BONIF_HATCHED_FILL if hatched else BONIF_SOLID_FILL
            _write_segment(1, start, end, label, fill)
    else:
        _write_segment(1, 1, num_days, "SP263", BONIF_SOLID_FILL)

    # Subvenciones (row 2)
    if subv_segments:
        for start, end, label, hatched in subv_segments:
            fill = SUBV_HATCHED_FILL if hatched else SUBV_SOLID_FILL
            _write_segment(2, start, end, label, fill)
    else:
        _write_segment(2, 1, num_days, "SP206", SUBV_SOLID_FILL)

    # FdR (row 3) - two sprints: days 1-15 and 16-last
    _write_segment(3, 1, 15, "FdR Sprint 1", FDR_SOLID_FILL)
    _write_segment(3, 16, num_days, "FdR Sprint 2", FDR_SOLID_FILL)

    # Transversal (row 4) - full month
    _write_segment(4, 1, num_days, "TRANSVERSAL", TRANSV_SOLID_FILL)

    return ws


class TestFindMesActualColumns:
    def test_finds_single_col(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=53, value="Mes Actual")
        assert _find_mes_actual_columns(ws) == [53]

    def test_excludes_gap_mes_actual(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=53, value="Mes Actual")
        ws.cell(row=DAY_LETTER_ROW, column=55, value="Gap Mes Actual")
        assert _find_mes_actual_columns(ws) == [53]

    def test_finds_two_cols(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=53, value="Mes Actual")
        ws.cell(row=DAY_LETTER_ROW, column=59, value="Mes Actual")
        assert _find_mes_actual_columns(ws) == [53, 59]


class TestFindLastBillableDay:
    def test_all_solid(self):
        from sprint_cert_automation.services.sprint_configurator import SprintSegment
        segments = [
            SprintSegment("SP1", 1, 1, 14, False, "bonificaciones"),
            SprintSegment("SP2", 2, 15, 28, False, "bonificaciones"),
        ]
        assert _find_last_billable_day(segments, "bonificaciones") == 28

    def test_last_is_hatched(self):
        from sprint_cert_automation.services.sprint_configurator import SprintSegment
        segments = [
            SprintSegment("SP1", 1, 1, 14, False, "bonificaciones"),
            SprintSegment("SP2", 2, 15, 25, False, "bonificaciones"),
            SprintSegment("SP3", 3, 26, 31, True, "bonificaciones"),
        ]
        assert _find_last_billable_day(segments, "bonificaciones") == 25

    def test_all_hatched(self):
        from sprint_cert_automation.services.sprint_configurator import SprintSegment
        segments = [
            SprintSegment("SP1", 1, 1, 31, True, "bonificaciones"),
        ]
        assert _find_last_billable_day(segments, "bonificaciones") is None

    def test_empty_segments(self):
        assert _find_last_billable_day([], "bonificaciones") is None


class TestUpdateRevenuesMesActualFormulas:
    def test_basic_formula_with_all_solid_sprints(self):
        """All sprints solid → SUM covers full month for all teams."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_oct", 2026, 10)

        updated = update_revenues_mes_actual_formulas(ws, 2026, 10, 31)
        assert updated == 31  # rows 7-37

        # Row 7: Transversal, rev_row=41, last day=31→col AQ
        f7 = ws.cell(row=7, column=53).value
        assert 'SUM(M41:AQ41)' in f7
        assert '"Transversal"' in f7

    def test_hatched_sprint_excluded_from_range(self):
        """Hatched sprints are excluded - SUM ends at last solid sprint."""
        wb = openpyxl.Workbook()
        # Bonif: solid days 1-25, hatched 26-31
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_nov", 2026, 11,
            bonif_segments=[
                (1, 14, "SP264", False),
                (15, 25, "SP265", False),
                (26, 30, "SP266", True),
            ])

        update_revenues_mes_actual_formulas(ws, 2026, 11, 30)

        # Bonif last solid day=25 → col M+24=col37=AK
        f8 = ws.cell(row=8, column=53).value
        assert '"Bonificaciones",SUM(M42:AK42)' in f8

        # Transversal still full month (day 30 → col AQ-1=AP? Let me compute)
        # FIRST_DAY_COL=13, day30→col42=AP
        f7 = ws.cell(row=7, column=53).value
        assert '"Transversal",SUM(M41:AP41)' in f7

    def test_subv_hatched_excluded(self):
        """Subvenciones hatched sprint excluded from billing range."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_nov", 2026, 11,
            subv_segments=[
                (1, 20, "SP207", False),
                (21, 30, "SP208", True),
            ])

        update_revenues_mes_actual_formulas(ws, 2026, 11, 30)

        # Subv last solid day=20 → col13+19=col32=AF
        f9 = ws.cell(row=9, column=53).value
        assert '"Subvenciones",SUM(M43:AF43)' in f9

    def test_fdr_always_full_month(self):
        """FdR always completes in-month, so SUM covers full month."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_oct", 2026, 10)

        update_revenues_mes_actual_formulas(ws, 2026, 10, 31)

        # FdR last day=31→AQ
        f10 = ws.cell(row=10, column=53).value
        assert '"Fondos de Reserva",SUM(M44:AQ44)' in f10

    def test_references_revenue_rows(self):
        """Formula references revenue rows (cost_row + 34)."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_oct", 2026, 10)

        update_revenues_mes_actual_formulas(ws, 2026, 10, 31)

        # Row 7 → rev row 41, Row 37 → rev row 71
        f7 = ws.cell(row=7, column=53).value
        assert "41" in f7
        f37 = ws.cell(row=37, column=53).value
        assert "71" in f37

    def test_team_column_in_formula(self):
        """Formula uses correct team column letter for IF checks."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_oct", 2026, 10)

        update_revenues_mes_actual_formulas(ws, 2026, 10, 31)

        # Team col is F (from Equipo header), references revenue row
        f7 = ws.cell(row=7, column=53).value
        assert f7.startswith('=IF(F41="Transversal"')

    def test_factura_column_formula(self):
        """Second Mes Actual column gets horas × tarifa formula."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_mes_actual(
            wb, "FY27_oct", 2026, 10,
            mes_actual_col=53, mes_actual_factura_col=59)

        updated = update_revenues_mes_actual_formulas(ws, 2026, 10, 31)
        # 31 horas + 31 factura = 62
        assert updated == 62

        # Factura: =BA7*G7
        f7_factura = ws.cell(row=7, column=59).value
        assert f7_factura == "=BA7*G7"

    def test_no_mes_actual_column_returns_zero(self):
        """Returns 0 when no Mes Actual columns found."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("test")

        updated = update_revenues_mes_actual_formulas(ws, 2026, 10, 31)
        assert updated == 0


# ---------------------------------------------------------------------------
# Revenues Periodo Actual No Facturable formula tests
# ---------------------------------------------------------------------------


def _create_sheet_with_sprints_and_gap_mes_actual(wb, name, year, month,
                                                   bonif_segments=None,
                                                   subv_segments=None,
                                                   gap_col=55,
                                                   gap_factura_col=None):
    """Create a sheet with sprint config and Gap Mes Actual column for testing."""
    from sprint_cert_automation.services.sprint_configurator import (
        BONIF_HATCHED_FILL,
        BONIF_SOLID_FILL,
        SUBV_HATCHED_FILL,
        SUBV_SOLID_FILL,
        FDR_SOLID_FILL,
        TRANSV_SOLID_FILL,
    )
    import calendar as cal_mod

    ws = wb.create_sheet(name)
    num_days = cal_mod.monthrange(year, month)[1]

    # Set up calendar row 5 (day numbers)
    for day in range(1, num_days + 1):
        ws.cell(row=DAY_NUMBER_ROW, column=FIRST_DAY_COL + day - 1, value=day)

    # Team column at F
    ws.cell(row=DAY_LETTER_ROW, column=6, value="Equipo")
    # Tarifa column at G
    ws.cell(row=DAY_LETTER_ROW, column=7, value="Tarifa")

    # Set team names for cost and revenue rows
    teams = ["Transversal", "Bonificaciones", "Subvenciones", "Fondos de Reserva"]
    for i, team in enumerate(teams):
        ws.cell(row=COST_FIRST_ROW + i, column=6, value=team)
        ws.cell(row=COST_FIRST_ROW + i, column=7, value=50)
        ws.cell(row=REVENUE_FIRST_ROW + i, column=6, value=team)

    # Gap Mes Actual header
    ws.cell(row=DAY_LETTER_ROW, column=gap_col, value="Gap Mes Actual")
    if gap_factura_col:
        ws.cell(row=DAY_LETTER_ROW, column=gap_factura_col, value="Gap Mes Actual")

    def _write_segment(row, start_day, end_day, label, fill):
        start_col = FIRST_DAY_COL + start_day - 1
        end_col = FIRST_DAY_COL + end_day - 1
        cell = ws.cell(row=row, column=start_col)
        cell.value = label
        cell.fill = fill
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=row, column=c).fill = fill
        if end_col > start_col:
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col)

    # Bonificaciones (row 1)
    if bonif_segments:
        for start, end, label, hatched in bonif_segments:
            fill = BONIF_HATCHED_FILL if hatched else BONIF_SOLID_FILL
            _write_segment(1, start, end, label, fill)
    else:
        _write_segment(1, 1, num_days, "SP263", BONIF_SOLID_FILL)

    # Subvenciones (row 2)
    if subv_segments:
        for start, end, label, hatched in subv_segments:
            fill = SUBV_HATCHED_FILL if hatched else SUBV_SOLID_FILL
            _write_segment(2, start, end, label, fill)
    else:
        _write_segment(2, 1, num_days, "SP206", SUBV_SOLID_FILL)

    # FdR (row 3) - always complete
    _write_segment(3, 1, 15, "FdR Sprint 1", FDR_SOLID_FILL)
    _write_segment(3, 16, num_days, "FdR Sprint 2", FDR_SOLID_FILL)

    # Transversal (row 4) - full month
    _write_segment(4, 1, num_days, "TRANSVERSAL", TRANSV_SOLID_FILL)

    return ws


class TestFindGapMesActualColumns:
    def test_finds_gap_mes_actual(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=55, value="Gap Mes Actual")
        assert _find_gap_mes_actual_columns(ws) == [55]

    def test_excludes_plain_mes_actual(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=DAY_LETTER_ROW, column=53, value="Mes Actual")
        ws.cell(row=DAY_LETTER_ROW, column=55, value="Gap Mes Actual")
        assert _find_gap_mes_actual_columns(ws) == [55]


class TestUpdateRevenuesNoFactFormulas:
    def test_all_sprints_solid_full_month(self):
        """All sprints cover full month → all formulas return 0."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_oct", 2026, 10)

        updated = update_revenues_no_fact_formulas(ws, 2026, 10, 31)
        assert updated == 31

        # All teams: 0 (full month covered)
        f7 = ws.cell(row=7, column=55).value
        assert '"Transversal",0' in f7
        assert '"Bonificaciones",0' in f7
        assert '"Subvenciones",0' in f7

    def test_hatched_sprint_range(self):
        """Hatched sprint days appear as non-billable SUM range."""
        wb = openpyxl.Workbook()
        # Bonif: solid 1-25, hatched 26-30
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_nov", 2026, 11,
            bonif_segments=[
                (1, 14, "SP264", False),
                (15, 25, "SP265", False),
                (26, 30, "SP266", True),
            ])

        update_revenues_no_fact_formulas(ws, 2026, 11, 30)

        # Bonif non-billable: day 26 to day 30
        # Day 26 → col 13+25=38=AL, Day 30 → col 13+29=42=AP
        f8 = ws.cell(row=8, column=55).value
        assert "SUM(AL42:AP42)" in f8

    def test_subv_hatched_range(self):
        """Subvenciones non-billable range computed correctly."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_nov", 2026, 11,
            subv_segments=[
                (1, 20, "SP207", False),
                (21, 30, "SP208", True),
            ])

        update_revenues_no_fact_formulas(ws, 2026, 11, 30)

        # Subv non-billable: day 21 to 30
        # Day 21 → col 13+20=33=AG, Day 30 → col 42=AP
        f9 = ws.cell(row=9, column=55).value
        assert "SUM(AG43:AP43)" in f9

    def test_transversal_always_zero(self):
        """Transversal always returns 0."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_nov", 2026, 11,
            bonif_segments=[(1, 25, "SP265", False), (26, 30, "SP266", True)])

        update_revenues_no_fact_formulas(ws, 2026, 11, 30)

        f7 = ws.cell(row=7, column=55).value
        assert '"Transversal",0' in f7

    def test_fdr_always_zero(self):
        """FdR always returns 0."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_nov", 2026, 11,
            bonif_segments=[(1, 25, "SP265", False), (26, 30, "SP266", True)])

        update_revenues_no_fact_formulas(ws, 2026, 11, 30)

        f10 = ws.cell(row=10, column=55).value
        assert '"Fondos de Reserva",0' in f10

    def test_references_revenue_rows(self):
        """Formula references revenue rows (row + 34)."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_nov", 2026, 11,
            bonif_segments=[(1, 25, "SP265", False), (26, 30, "SP266", True)])

        update_revenues_no_fact_formulas(ws, 2026, 11, 30)

        # Row 7 → rev 41, Row 37 → rev 71
        f7 = ws.cell(row=7, column=55).value
        assert "41" in f7
        f37 = ws.cell(row=37, column=55).value
        assert "71" in f37

    def test_single_day_no_sum(self):
        """Single non-billable day uses direct cell ref."""
        wb = openpyxl.Workbook()
        # Bonif solid 1-30 in a 31-day month → day 31 non-billable
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_oct", 2026, 10,
            bonif_segments=[(1, 30, "SP265", False)])
        # Not hatched, but sprint ends day 30 in a 31-day month

        update_revenues_no_fact_formulas(ws, 2026, 10, 31)

        # Day 31 → col 43=AQ, single cell
        f8 = ws.cell(row=8, column=55).value
        assert "AQ42" in f8
        assert "SUM" not in f8.split("Bonificaciones")[1].split(",")[0]

    def test_factura_column(self):
        """Second Gap Mes Actual column gets horas × tarifa."""
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_sprints_and_gap_mes_actual(
            wb, "FY27_oct", 2026, 10,
            gap_col=55, gap_factura_col=61)

        updated = update_revenues_no_fact_formulas(ws, 2026, 10, 31)
        assert updated == 62  # 31 + 31

        f7_factura = ws.cell(row=7, column=61).value
        assert f7_factura == "=BC7*G7"

    def test_no_gap_columns_returns_zero(self):
        """Returns 0 when no Gap Mes Actual columns found."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("test")
        assert update_revenues_no_fact_formulas(ws, 2026, 10, 31) == 0
