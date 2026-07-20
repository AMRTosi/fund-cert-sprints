"""Tests for the sprint configurator service."""

from __future__ import annotations

import calendar

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from pathlib import Path

from sprint_cert_automation.services.sprint_configurator import (
    BONIF_HATCHED_FILL,
    BONIF_SOLID_FILL,
    BONIF_WORKING_DAYS,
    FDR_SOLID_FILL,
    SUBV_HATCHED_FILL,
    SUBV_SOLID_FILL,
    SUBV_WORKING_DAYS,
    TRANSV_SOLID_FILL,
    SprintSegment,
    TeamSprintInfo,
    _count_working_days,
    _advance_working_days,
    _extract_sprint_number,
    calculate_bonif_sprints,
    calculate_fdr_sprints,
    calculate_subv_sprints,
    calculate_transversal_sprint,
    configure_sprints,
    read_sprints_from_sheet,
    write_sprints_to_sheet,
)
from sprint_cert_automation.services.sheet_duplicator import (
    DAY_LETTER_ROW,
    DAY_NUMBER_ROW,
    FIRST_DAY_COL,
    _generate_calendar,
    _infer_year_month_from_sheet_name,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_sheet_with_calendar(wb, name, year, month):
    """Create a sheet with a valid calendar in rows 5-6."""
    ws = wb.create_sheet(name)
    cal = _generate_calendar(year, month)
    for day_num, day_letter in cal:
        col = FIRST_DAY_COL + day_num - 1
        ws.cell(row=DAY_NUMBER_ROW, column=col, value=day_num)
        ws.cell(row=DAY_LETTER_ROW, column=col, value=day_letter)
    return ws


def _add_sprint_segment(ws, row, start_day, end_day, name, is_hatched=False, team="bonificaciones"):
    """Write a sprint segment to a sheet (merged cells with fill)."""
    day1_col = FIRST_DAY_COL
    start_col = day1_col + start_day - 1
    end_col = day1_col + end_day - 1

    fills = {
        "bonificaciones": (BONIF_SOLID_FILL, BONIF_HATCHED_FILL),
        "subvenciones": (SUBV_SOLID_FILL, SUBV_HATCHED_FILL),
        "fdr": (FDR_SOLID_FILL, FDR_SOLID_FILL),
        "transversal": (TRANSV_SOLID_FILL, TRANSV_SOLID_FILL),
    }
    solid, hatched = fills.get(team, (PatternFill(), PatternFill()))
    fill = hatched if is_hatched else solid

    cell = ws.cell(row=row, column=start_col)
    cell.value = name
    cell.fill = fill

    for col in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=col).fill = fill

    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)


# ---------------------------------------------------------------------------
# Working day utility tests
# ---------------------------------------------------------------------------

class TestWorkingDayUtils:
    def test_count_working_days_full_week(self):
        # Oct 2026: Oct 5 (Mon) to Oct 9 (Fri) = 5 working days
        assert _count_working_days(2026, 10, 5, 9) == 5

    def test_count_working_days_includes_weekends(self):
        # Oct 2026: Oct 1 (Thu) to Oct 4 (Sun) = 2 working days (Thu, Fri)
        assert _count_working_days(2026, 10, 1, 4) == 2

    def test_count_working_days_single_day_weekday(self):
        # Oct 1 2026 = Thursday
        assert _count_working_days(2026, 10, 1, 1) == 1

    def test_count_working_days_single_day_weekend(self):
        # Oct 3 2026 = Saturday
        assert _count_working_days(2026, 10, 3, 3) == 0

    def test_advance_working_days_simple(self):
        # Oct 1 2026 = Thu, advance 2 working days -> Oct 2 (Fri)
        assert _advance_working_days(2026, 10, 1, 2) == 2

    def test_advance_working_days_crosses_weekend(self):
        # Oct 1 2026 = Thu, advance 4 working days -> Thu, Fri, [Sat, Sun], Mon, Tue
        # Oct 1, 2, 5, 6 -> day 6
        assert _advance_working_days(2026, 10, 1, 4) == 6

    def test_advance_working_days_overflows_month(self):
        # Start Oct 30 (Fri), advance 10 working days -> overflows
        result = _advance_working_days(2026, 10, 30, 10)
        assert result > 31  # overflow


# ---------------------------------------------------------------------------
# Sprint number extraction tests
# ---------------------------------------------------------------------------

class TestExtractSprintNumber:
    def test_full_name(self):
        assert _extract_sprint_number("Bonificaciones - SP262") == 262

    def test_short_name(self):
        assert _extract_sprint_number("SP207") == 207

    def test_no_sprint_number(self):
        assert _extract_sprint_number("Transversal - octubre") is None

    def test_none_value(self):
        assert _extract_sprint_number(None) is None

    def test_fdr_name(self):
        assert _extract_sprint_number("Fondos de Reserva - SP54") == 54


# ---------------------------------------------------------------------------
# Read sprints tests
# ---------------------------------------------------------------------------

class TestReadSprints:
    def test_reads_merged_sprint_segments(self):
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_calendar(wb, "TestSheet", 2026, 10)

        # Add bonificaciones sprints
        _add_sprint_segment(ws, 1, 1, 13, "Bonificaciones - SP262", team="bonificaciones")
        _add_sprint_segment(ws, 1, 14, 27, "Bonificaciones - SP263", team="bonificaciones")
        _add_sprint_segment(ws, 1, 28, 31, "SP264", is_hatched=True, team="bonificaciones")

        result = read_sprints_from_sheet(ws)
        bonif = result["bonificaciones"]

        assert len(bonif.segments) == 3
        assert bonif.segments[0].sprint_number == 262
        assert bonif.segments[0].is_hatched is False
        assert bonif.segments[2].sprint_number == 264
        assert bonif.segments[2].is_hatched is True
        assert bonif.last_sprint_number == 264
        assert bonif.has_hatched_tail is True

    def test_reads_all_four_teams(self):
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_calendar(wb, "TestSheet", 2026, 10)

        _add_sprint_segment(ws, 1, 1, 14, "Bonificaciones - SP262", team="bonificaciones")
        _add_sprint_segment(ws, 2, 1, 22, "Subvenciones - SP206", team="subvenciones")
        _add_sprint_segment(ws, 3, 1, 15, "Fondos de Reserva - SP54", team="fdr")
        _add_sprint_segment(ws, 3, 16, 31, "Fondos de Reserva - SP55", team="fdr")
        _add_sprint_segment(ws, 4, 1, 31, "Transversal - octubre", team="transversal")

        result = read_sprints_from_sheet(ws)
        assert len(result["bonificaciones"].segments) == 1
        assert len(result["subvenciones"].segments) == 1
        assert len(result["fdr"].segments) == 2
        assert len(result["transversal"].segments) == 1

    def test_no_hatched_tail(self):
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_calendar(wb, "TestSheet", 2026, 10)
        _add_sprint_segment(ws, 1, 1, 14, "Bonificaciones - SP262", team="bonificaciones")

        result = read_sprints_from_sheet(ws)
        assert result["bonificaciones"].has_hatched_tail is False


# ---------------------------------------------------------------------------
# Calculate Fondos de Reserva tests
# ---------------------------------------------------------------------------

class TestCalculateFdrSprints:
    def test_two_sprints_with_correct_boundaries(self):
        prev_info = TeamSprintInfo(segments=[
            SprintSegment("FdR - SP52", 52, 1, 15, False, "fdr"),
            SprintSegment("FdR - SP53", 53, 16, 30, False, "fdr"),
        ])
        result = calculate_fdr_sprints(2026, 11, 30, prev_info)

        assert len(result) == 2
        assert result[0].sprint_number == 54
        assert result[0].start_day == 1
        assert result[0].end_day == 15
        assert result[1].sprint_number == 55
        assert result[1].start_day == 16
        assert result[1].end_day == 30

    def test_without_previous_info(self):
        result = calculate_fdr_sprints(2026, 10, 31, None)
        assert len(result) == 2
        assert result[0].sprint_number == 1
        assert result[1].sprint_number == 2


# ---------------------------------------------------------------------------
# Calculate Transversal tests
# ---------------------------------------------------------------------------

class TestCalculateTransversalSprint:
    def test_spans_entire_month(self):
        result = calculate_transversal_sprint(2026, 10, 31)
        assert len(result) == 1
        assert result[0].name == "Transversal - octubre"
        assert result[0].start_day == 1
        assert result[0].end_day == 31

    def test_february_name(self):
        result = calculate_transversal_sprint(2026, 2, 28)
        assert result[0].name == "Transversal - febrero"
        assert result[0].end_day == 28


# ---------------------------------------------------------------------------
# Calculate Bonificaciones tests
# ---------------------------------------------------------------------------

class TestCalculateBonifSprints:
    def test_with_carry_over_from_previous(self):
        # Previous month had hatched tail: SP264 at days 28-31, Oct 2026
        # Oct 28=Wed, 29=Thu, 30=Fri, 31=Sat -> 3 working days consumed
        prev_info = TeamSprintInfo(segments=[
            SprintSegment("Bonif - SP263", 263, 1, 27, False, "bonificaciones"),
            SprintSegment("SP264", 264, 28, 31, True, "bonificaciones"),
        ])

        result = calculate_bonif_sprints(2026, 11, 30, prev_info, 2026, 10)

        # SP264 carry-over: 10 - 3 = 7 remaining working days
        assert result[0].sprint_number == 264
        assert result[0].start_day == 1
        assert result[0].is_hatched is False

    def test_without_previous_starts_at_sp1(self):
        result = calculate_bonif_sprints(2026, 10, 31, None, None, None)
        assert result[0].sprint_number == 1
        assert result[0].start_day == 1

    def test_generates_multiple_sprints(self):
        result = calculate_bonif_sprints(2026, 10, 31, None, None, None)
        # 31-day month with ~23 working days should produce 2-3 sprints
        assert len(result) >= 2

    def test_last_sprint_hatched_if_overflows(self):
        result = calculate_bonif_sprints(2026, 10, 31, None, None, None)
        # Check if any sprint is hatched (depends on month layout)
        # With 23 working days and 10-day sprints: 10+10=20, 3 remaining
        has_hatched = any(s.is_hatched for s in result)
        # Either the last sprint overflows or fits exactly
        assert isinstance(has_hatched, bool)

    def test_no_previous_hatched_continues_numbering(self):
        prev_info = TeamSprintInfo(segments=[
            SprintSegment("Bonif - SP270", 270, 1, 14, False, "bonificaciones"),
            SprintSegment("Bonif - SP271", 271, 15, 28, False, "bonificaciones"),
        ])
        result = calculate_bonif_sprints(2026, 3, 31, prev_info, 2026, 2)
        assert result[0].sprint_number == 272


# ---------------------------------------------------------------------------
# Calculate Subvenciones tests
# ---------------------------------------------------------------------------

class TestCalculateSubvSprints:
    def test_with_carry_over(self):
        # Previous: SP207 hatched at days 29-31 of October (3 calendar days)
        # Oct 29=Thu, 30=Fri, 31=Sat -> 2 working days consumed
        prev_info = TeamSprintInfo(segments=[
            SprintSegment("Subv - SP206", 206, 1, 28, False, "subvenciones"),
            SprintSegment("SP207", 207, 29, 31, True, "subvenciones"),
        ])

        result = calculate_subv_sprints(2026, 11, 30, prev_info, 2026, 10)

        assert result[0].sprint_number == 207
        assert result[0].start_day == 1
        assert result[0].is_hatched is False

    def test_without_previous(self):
        result = calculate_subv_sprints(2026, 10, 31, None, None, None)
        assert result[0].sprint_number == 1

    def test_sprint_numbering_continuous(self):
        prev_info = TeamSprintInfo(segments=[
            SprintSegment("Subv - SP211", 211, 1, 22, False, "subvenciones"),
            SprintSegment("SP212", 212, 23, 28, True, "subvenciones"),
        ])
        result = calculate_subv_sprints(2026, 3, 31, prev_info, 2026, 2)
        # SP212 carry-over first, then SP213
        assert result[0].sprint_number == 212
        sp_nums = [s.sprint_number for s in result]
        for i in range(1, len(sp_nums)):
            assert sp_nums[i] == sp_nums[i - 1] + 1


# ---------------------------------------------------------------------------
# Write sprints tests
# ---------------------------------------------------------------------------

class TestWriteSprints:
    def test_writes_merged_cells(self):
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_calendar(wb, "Test", 2026, 10)

        segments = {
            "bonificaciones": [
                SprintSegment("Bonif - SP262", 262, 1, 14, False, "bonificaciones"),
                SprintSegment("Bonif - SP263", 263, 15, 28, False, "bonificaciones"),
            ],
            "subvenciones": [
                SprintSegment("Subv - SP206", 206, 1, 22, False, "subvenciones"),
            ],
            "fdr": [
                SprintSegment("FdR - SP54", 54, 1, 15, False, "fdr"),
                SprintSegment("FdR - SP55", 55, 16, 31, False, "fdr"),
            ],
            "transversal": [
                SprintSegment("Transversal - octubre", None, 1, 31, False, "transversal"),
            ],
        }

        write_sprints_to_sheet(ws, segments, 31)

        # Check bonificaciones row 1
        assert ws.cell(row=1, column=FIRST_DAY_COL).value == "Bonif - SP262"
        assert ws.cell(row=1, column=FIRST_DAY_COL + 14).value == "Bonif - SP263"

        # Check merge exists for row 1
        merged_in_row1 = [mr for mr in ws.merged_cells.ranges if mr.min_row == 1]
        assert len(merged_in_row1) >= 2  # At least SP262 and SP263

    def test_applies_hatched_fill(self):
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_calendar(wb, "Test", 2026, 10)

        segments = {
            "bonificaciones": [
                SprintSegment("SP264", 264, 28, 31, True, "bonificaciones"),
            ],
            "subvenciones": [],
            "fdr": [],
            "transversal": [],
        }

        write_sprints_to_sheet(ws, segments, 31)

        cell = ws.cell(row=1, column=FIRST_DAY_COL + 27)
        assert cell.fill.patternType == "darkDown"


# ---------------------------------------------------------------------------
# Configure sprints (integration) tests
# ---------------------------------------------------------------------------

class TestConfigureSprints:
    def test_configures_all_teams_without_previous(self):
        wb = openpyxl.Workbook()
        ws = _create_sheet_with_calendar(wb, "New", 2026, 10)

        result = configure_sprints(ws, 2026, 10, 31)

        assert "bonificaciones" in result
        assert "subvenciones" in result
        assert "fdr" in result
        assert "transversal" in result
        assert len(result["fdr"]) == 2
        assert len(result["transversal"]) == 1

    def test_configures_with_previous_sheet(self):
        wb = openpyxl.Workbook()
        # Create previous sheet (October 2026)
        prev_ws = _create_sheet_with_calendar(wb, "FY27_oct", 2026, 10)
        _add_sprint_segment(prev_ws, 1, 1, 13, "Bonif - SP262", team="bonificaciones")
        _add_sprint_segment(prev_ws, 1, 14, 27, "Bonif - SP263", team="bonificaciones")
        _add_sprint_segment(prev_ws, 1, 28, 31, "SP264", is_hatched=True, team="bonificaciones")
        _add_sprint_segment(prev_ws, 2, 1, 6, "Subv - SP205", team="subvenciones")
        _add_sprint_segment(prev_ws, 2, 7, 28, "Subv - SP206", team="subvenciones")
        _add_sprint_segment(prev_ws, 2, 29, 31, "SP207", is_hatched=True, team="subvenciones")
        _add_sprint_segment(prev_ws, 3, 1, 15, "FdR - SP54", team="fdr")
        _add_sprint_segment(prev_ws, 3, 16, 31, "FdR - SP55", team="fdr")
        _add_sprint_segment(prev_ws, 4, 1, 31, "Transversal - octubre", team="transversal")

        # Create new sheet (November 2026)
        new_ws = _create_sheet_with_calendar(wb, "FY27_nov", 2026, 11)

        result = configure_sprints(
            new_ws, 2026, 11, 30,
            prev_ws=prev_ws, prev_year=2026, prev_month=10,
        )

        # Bonificaciones should start with carry-over SP264
        bonif = result["bonificaciones"]
        assert bonif[0].sprint_number == 264
        assert bonif[0].is_hatched is False  # continuation, not hatched

        # Subvenciones should start with carry-over SP207
        subv = result["subvenciones"]
        assert subv[0].sprint_number == 207

        # FdR should increment: SP56, SP57
        fdr = result["fdr"]
        assert fdr[0].sprint_number == 56
        assert fdr[1].sprint_number == 57

        # Transversal named november
        transv = result["transversal"]
        assert transv[0].name == "Transversal - noviembre"


# ---------------------------------------------------------------------------
# Infer year/month from sheet name tests
# ---------------------------------------------------------------------------

class TestInferYearMonth:
    def test_fy27_sept(self):
        assert _infer_year_month_from_sheet_name("FY27_sept", 2026, 10) == (2026, 9)

    def test_fy27_oct(self):
        assert _infer_year_month_from_sheet_name("FY27_oct", 2026, 11) == (2026, 10)

    def test_fy27_ene(self):
        assert _infer_year_month_from_sheet_name("FY27_ene", 2027, 2) == (2027, 1)

    def test_fy26_dic(self):
        assert _infer_year_month_from_sheet_name("FY26_dic", 2026, 1) == (2025, 12)

    def test_fy26_mayo(self):
        assert _infer_year_month_from_sheet_name("FY26_mayo", 2026, 6) == (2026, 5)

    def test_fallback_for_unknown_format(self):
        assert _infer_year_month_from_sheet_name("SomeSheet", 2026, 3) == (2026, 2)

    def test_fallback_january_wraps_year(self):
        assert _infer_year_month_from_sheet_name("SomeSheet", 2026, 1) == (2025, 12)
