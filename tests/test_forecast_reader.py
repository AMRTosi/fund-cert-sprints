from datetime import date

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from sprint_cert_automation.services.forecast_reader import ForecastReader


def test_reader_merges_previous_and_current_month_segments(tmp_path) -> None:
    workbook_path = tmp_path / "forecast.xlsx"
    workbook = Workbook()
    may_sheet = workbook.active
    may_sheet.title = "FY26_mayo"
    june_sheet = workbook.create_sheet("FY26_jun")

    _populate_day_row(may_sheet, start_column=12, days=31)
    _populate_day_row(june_sheet, start_column=13, days=30)

    hatched_fill = PatternFill(patternType="darkTrellis")

    may_sheet.merge_cells(start_row=1, start_column=30, end_row=1, end_column=42)
    may_sheet.cell(row=1, column=30).value = "Bonificaciones - SP277"
    may_sheet.cell(row=1, column=30).fill = hatched_fill

    june_sheet.cell(row=1, column=13).value = "SP277"
    june_sheet.merge_cells(start_row=1, start_column=14, end_row=1, end_column=27)
    june_sheet.cell(row=1, column=14).value = "Bonificaciones - SP278"

    workbook.save(workbook_path)

    result = ForecastReader(workbook_path).read_sprints_for_target_window(2026, 6)
    by_id = {sprint.sprint_id: sprint for sprint in result}

    sprint_277 = by_id["SP277"]
    assert sprint_277.team == "Bonificaciones"
    assert sprint_277.start_date == date(2026, 5, 19)
    assert sprint_277.end_date == date(2026, 6, 1)
    assert not sprint_277.is_hatched
    assert sprint_277.source_sheet == "FY26_jun"

    sprint_278 = by_id["SP278"]
    assert sprint_278.start_date == date(2026, 6, 2)
    assert sprint_278.end_date == date(2026, 6, 15)


def test_reader_marks_hatched_current_month_sprints_and_transversal_ids(tmp_path) -> None:
    workbook_path = tmp_path / "forecast.xlsx"
    workbook = Workbook()
    may_sheet = workbook.active
    may_sheet.title = "FY26_mayo"
    june_sheet = workbook.create_sheet("FY26_jun")

    _populate_day_row(may_sheet, start_column=12, days=31)
    _populate_day_row(june_sheet, start_column=13, days=30)

    hatched_fill = PatternFill(patternType="darkTrellis")

    june_sheet.merge_cells(start_row=2, start_column=37, end_row=2, end_column=42)
    june_sheet.cell(row=2, column=37).value = "Subvenciones - SP217"
    june_sheet.cell(row=2, column=37).fill = hatched_fill

    june_sheet.merge_cells(start_row=4, start_column=13, end_row=4, end_column=42)
    june_sheet.cell(row=4, column=13).value = "Transversal - junio"

    workbook.save(workbook_path)

    result = ForecastReader(workbook_path).read_sprints_for_target_window(2026, 6)
    by_key = {(sprint.team, sprint.sprint_id): sprint for sprint in result}

    sprint_217 = by_key[("Subvenciones", "SP217")]
    assert sprint_217.is_hatched
    assert sprint_217.start_date == date(2026, 6, 25)
    assert sprint_217.end_date == date(2026, 6, 30)

    transversal = by_key[("AccentureTransversal", "2026-06")]
    assert transversal.start_date == date(2026, 6, 1)
    assert transversal.end_date == date(2026, 6, 30)


def test_reader_builds_real_holidays_and_workloads_for_cross_month_sprint(tmp_path) -> None:
    workbook_path = tmp_path / "forecast.xlsx"
    workbook = Workbook()
    may_sheet = workbook.active
    may_sheet.title = "FY26_mayo"
    june_sheet = workbook.create_sheet("FY26_jun")

    _populate_headers_may(may_sheet)
    _populate_headers_june(june_sheet)
    _populate_day_row(may_sheet, start_column=12, days=31)
    _populate_day_row(june_sheet, start_column=13, days=30)

    holiday_fill = PatternFill(patternType="solid", fgColor="FFFF00")
    grey_fill = PatternFill(patternType="solid", fgColor="D0D0D0")

    may_sheet.merge_cells(start_row=1, start_column=40, end_row=1, end_column=42)
    may_sheet.cell(row=1, column=40).value = "Bonificaciones - SP277"
    june_sheet.merge_cells(start_row=1, start_column=13, end_row=1, end_column=14)
    june_sheet.cell(row=1, column=13).value = "SP277"

    may_sheet.cell(row=5, column=40).fill = holiday_fill
    june_sheet.cell(row=5, column=13).fill = holiday_fill

    may_sheet.cell(row=7, column=2).value = "Ana"
    may_sheet.cell(row=7, column=5).value = "Bonificaciones"
    may_sheet.cell(row=7, column=40).fill = grey_fill

    june_sheet.cell(row=7, column=2).value = "Ana"
    june_sheet.cell(row=7, column=3).value = "Ana"
    june_sheet.cell(row=7, column=4).value = "Equipo Desarrollo"
    june_sheet.cell(row=7, column=5).value = "Consultor"
    june_sheet.cell(row=7, column=6).value = "Bonificaciones"
    june_sheet.cell(row=7, column=13).fill = grey_fill

    june_sheet.cell(row=8, column=2).value = "Luis"
    june_sheet.cell(row=8, column=3).value = "Luis"
    june_sheet.cell(row=8, column=4).value = "Equipo Desarrollo"
    june_sheet.cell(row=8, column=5).value = "Analista"
    june_sheet.cell(row=8, column=6).value = "Bonificaciones"

    workbook.save(workbook_path)

    reader = ForecastReader(workbook_path)
    sprint = next(
        item for item in reader.read_sprints_for_target_window(2026, 6) if item.sprint_id == "SP277"
    )
    holidays, workloads = reader.read_draft_data_for_sprint(sprint)

    assert [holiday.holiday_date for holiday in holidays] == [date(2026, 5, 29), date(2026, 6, 1)]
    assert all(holiday.label == "Festivo" for holiday in holidays)

    by_name = {workload.member.name: workload for workload in workloads}
    assert by_name["Ana"].member.billing_line == "Equipo Desarrollo"
    assert by_name["Ana"].member.category == "Consultor"
    assert by_name["Ana"].sprint_hours == 8.5
    assert by_name["Ana"].free_hours == 0.0
    assert by_name["Luis"].member.category == "Analista"
    assert by_name["Luis"].free_hours == 0.0


def test_reader_uses_summer_day_hours_between_june_15_and_september_15(tmp_path) -> None:
    workbook_path = tmp_path / "forecast.xlsx"
    workbook = Workbook()
    may_sheet = workbook.active
    may_sheet.title = "FY26_mayo"
    june_sheet = workbook.create_sheet("FY26_jun")

    _populate_headers_may(may_sheet)
    _populate_headers_june(june_sheet)
    _populate_day_row(may_sheet, start_column=12, days=31)
    _populate_day_row(june_sheet, start_column=13, days=30)

    june_sheet.merge_cells(start_row=1, start_column=28, end_row=1, end_column=30)
    june_sheet.cell(row=1, column=28).value = "Bonificaciones - SP300"

    june_sheet.cell(row=7, column=2).value = "Ana"
    june_sheet.cell(row=7, column=3).value = "Ana"
    june_sheet.cell(row=7, column=4).value = "Equipo Desarrollo"
    june_sheet.cell(row=7, column=5).value = "Consultor"
    june_sheet.cell(row=7, column=6).value = "Bonificaciones"

    workbook.save(workbook_path)

    reader = ForecastReader(workbook_path)
    sprint = next(
        item for item in reader.read_sprints_for_target_window(2026, 6) if item.sprint_id == "SP300"
    )
    _, workloads = reader.read_draft_data_for_sprint(sprint)

    assert len(workloads) == 1
    assert workloads[0].sprint_hours == 22.5


def _populate_day_row(worksheet, start_column: int, days: int) -> None:
    for offset in range(days):
        worksheet.cell(row=5, column=start_column + offset).value = offset + 1


def _populate_headers_may(worksheet) -> None:
    worksheet.cell(row=6, column=2).value = "Nombre y Apellidos"
    worksheet.cell(row=6, column=4).value = "Perfil Facturable"
    worksheet.cell(row=6, column=5).value = "Equipo"


def _populate_headers_june(worksheet) -> None:
    worksheet.cell(row=6, column=2).value = "Nombre y Apellidos"
    worksheet.cell(row=6, column=3).value = "Técnico"
    worksheet.cell(row=6, column=4).value = "Facturación"
    worksheet.cell(row=6, column=5).value = "Perfil Facturable"
    worksheet.cell(row=6, column=6).value = "Equipo"