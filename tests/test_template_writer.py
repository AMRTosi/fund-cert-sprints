from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.datetime import CALENDAR_MAC_1904, CALENDAR_WINDOWS_1900

from sprint_cert_automation.domain.models import CertificateDraft, Holiday, TeamMember, TeamMemberWorkload
from sprint_cert_automation.services.template_writer import TemplateWriter


def test_template_writer_populates_config_holidays_and_workloads(tmp_path) -> None:
    template_path = tmp_path / "template.xlsm"
    output_path = tmp_path / "output.xlsm"
    _build_template(template_path)

    draft = CertificateDraft(
        team="AccentureTransversal",
        sprint_id="2026-06",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 30),
        file_name="unused.xlsm",
        product_label="AccentureTransversal",
        holidays=[
            Holiday(holiday_date=date(2026, 6, 15), label="San Isidro"),
            Holiday(holiday_date=date(2026, 6, 24), label="San Juan"),
        ],
        workloads=[
            TeamMemberWorkload(
                member=TeamMember(
                    name="Ana",
                    billing_line="L1",
                    category="CONSULTOR JUNIOR",
                    team="AccentureTransversal",
                ),
                sprint_hours=170.0,
                free_hours=8.5,
            ),
            TeamMemberWorkload(
                member=TeamMember(
                    name="Luis",
                    billing_line="L2",
                    category="ANALISTA BI",
                    team="AccentureTransversal",
                ),
                sprint_hours=161.5,
                free_hours=17.0,
            ),
            TeamMemberWorkload(
                member=TeamMember(
                    name="Ana",
                    billing_line="L1-duplicado",
                    category="ANALISTA BI",
                    team="AccentureTransversal",
                ),
                sprint_hours=170.0,
                free_hours=0.0,
            ),
        ],
    )

    TemplateWriter(template_path).write(draft, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Config"]
    try:
        assert worksheet["A2"].value.date() == date(2026, 6, 1)
        assert worksheet["B2"].value.date() == date(2026, 6, 30)
        assert worksheet["D2"].value == "Junio"
        assert worksheet["G3"].value == "Transversales"

        assert worksheet["A12"].value == "San Isidro"
        assert worksheet["B12"].value.date() == date(2026, 6, 15)
        assert worksheet["A13"].value == "San Juan"
        assert worksheet["B13"].value.date() == date(2026, 6, 24)

        assert worksheet["F13"].value == "Ana"
        assert worksheet["G13"].value == "L1"
        assert worksheet["H13"].value == "Consultor Junior"
        assert worksheet["I13"].value == 170.0
        assert worksheet["J13"].value == 8.5

        assert worksheet["F14"].value == "Luis"
        assert worksheet["G14"].value == "L2"
        assert worksheet["H14"].value == "Analista BI"
        assert worksheet["I14"].value == 161.5
        assert worksheet["J14"].value == 17.0
        assert worksheet["F15"].value is None
    finally:
        workbook.close()


def test_template_writer_expands_tables_when_needed(tmp_path) -> None:
    template_path = tmp_path / "template.xlsm"
    output_path = tmp_path / "output.xlsm"
    _build_template(template_path, holiday_rows=1, team_rows=1)

    draft = CertificateDraft(
        team="Fondos de Reserva MRR",
        sprint_id="SP70",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 15),
        file_name="unused.xlsm",
        product_label="Fondos de Reserva MRR",
        holidays=[
            Holiday(holiday_date=date(2026, 6, 2), label="F1"),
            Holiday(holiday_date=date(2026, 6, 3), label="F2"),
        ],
        workloads=[
            TeamMemberWorkload(
                member=TeamMember("A", "B1", "Product Owner", "Fondos de Reserva MRR"),
                sprint_hours=10.0,
                free_hours=0.0,
            ),
            TeamMemberWorkload(
                member=TeamMember("B", "B2", "Consultor Junior", "Fondos de Reserva MRR"),
                sprint_hours=11.0,
                free_hours=1.0,
            ),
            TeamMemberWorkload(
                member=TeamMember("C", "B3", "Analista BI", "Fondos de Reserva MRR"),
                sprint_hours=12.0,
                free_hours=2.0,
            ),
        ],
    )

    TemplateWriter(template_path).write(draft, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Config"]
    try:
        assert worksheet.tables["TB_Festivos"].ref == "A11:B13"
        assert worksheet.tables["TB_Equipo"].ref == "F12:J15"
        assert worksheet["G3"].value == "Fondos de Reserva"
        assert worksheet["D2"].value == "SP70"
    finally:
        workbook.close()


def test_template_writer_converts_com_dates_to_excel_serials(tmp_path) -> None:
    writer = TemplateWriter(tmp_path / "template.xlsm")

    assert writer._to_excel_serial_date(date(2026, 7, 1), CALENDAR_WINDOWS_1900) == 46204.0
    assert writer._to_excel_serial_date(date(2026, 7, 1), CALENDAR_MAC_1904) == 44742.0


def test_template_writer_writes_com_config_as_serial_dates(tmp_path) -> None:
    writer = TemplateWriter(tmp_path / "template.xlsm")
    draft = CertificateDraft(
        team="Bonificaciones",
        sprint_id="SP280",
        start_date=date(2026, 7, 1),
        end_date=date(2026, 7, 20),
        file_name="unused.xlsm",
        product_label="Bonificaciones",
    )
    worksheet = _FakeComWorksheet(date1904=False)

    writer._write_config_com(worksheet, draft)

    assert worksheet.ranges["A2"].Value2 == 46204.0
    assert worksheet.ranges["B2"].Value2 == 46223.0
    assert worksheet.ranges["D2"].Value == "SP280"
    assert worksheet.ranges["G3"].Value == "Bonificaciones"


def _build_template(path, holiday_rows: int = 8, team_rows: int = 4) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Config"
    master = workbook.create_sheet("Maestra")

    worksheet["A1"] = "Fecha Inicio"
    worksheet["B1"] = "Fecha Fin"
    worksheet["D1"] = "Nº"
    worksheet["G3"] = "Transversales"

    worksheet["A11"] = "Festivo"
    worksheet["B11"] = "Día"
    worksheet["F12"] = "Técnico"
    worksheet["G12"] = "Facturación"
    worksheet["H12"] = "Categoría"
    worksheet["I12"] = "Horas Sprint"
    worksheet["J12"] = "Horas libres"

    for row in range(12, 12 + holiday_rows):
        worksheet.cell(row=row, column=1).fill = PatternFill(patternType="solid", fgColor="DDDDDD")
        worksheet.cell(row=row, column=2).fill = PatternFill(patternType="solid", fgColor="DDDDDD")

    for row in range(13, 13 + team_rows):
        for column in range(6, 11):
            worksheet.cell(row=row, column=column).fill = PatternFill(patternType="solid", fgColor="EEEEEE")

    holiday_table = Table(displayName="TB_Festivos", ref=f"A11:B{11 + holiday_rows}")
    team_table = Table(displayName="TB_Equipo", ref=f"F12:J{12 + team_rows}")
    profile_table = Table(displayName="TB_Perfiles", ref="G1:G4")
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    holiday_table.tableStyleInfo = style
    team_table.tableStyleInfo = style
    profile_table.tableStyleInfo = style
    worksheet.add_table(holiday_table)
    worksheet.add_table(team_table)

    master["G1"] = "Categoria"
    master["G2"] = "Product Owner Proxy"
    master["G3"] = "Consultor Junior"
    master["G4"] = "Analista BI"
    master.add_table(profile_table)

    workbook.save(path)
    workbook.close()


class _FakeComWorksheet:
    def __init__(self, date1904: bool) -> None:
        self.Parent = _FakeComWorkbook(date1904)
        self.ranges = {
            "A2": _FakeComRange(),
            "B2": _FakeComRange(),
            "D2": _FakeComRange(),
            "G3": _FakeComRange(),
        }

    def Range(self, address: str) -> "_FakeComRange":
        return self.ranges[address]


class _FakeComWorkbook:
    def __init__(self, date1904: bool) -> None:
        self.Date1904 = date1904


class _FakeComRange:
    def __init__(self) -> None:
        self.Value = None
        self.Value2 = None
        self.NumberFormat = None